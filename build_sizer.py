#!/usr/bin/env python3
"""
A&S Capital Sizer -- Excel Template Builder  (v4 - Eastview Architecture)
Replicates Eastview's proven sizer structure with Colchis pricing/leverage.

Key design decisions (matching Eastview):
  - Borrower Classification (A+/A/B) computed from experience
  - Leverage lookup by "LoanPurpose / Classification / FICO"
  - Core sizing: MIN(Front-End, Back-End) with product-specific logic
  - Front End = AsIs × LTV  (and LTC × PurchasePrice for Purchase)
  - Back End  = ARV × LTARV  (and TLTC × TotalCost for Purchase)
  - Initial Loan Amount sized first; rehab + IR layered on top
  - Interest Reserves = 5% of (LTARV × ARV) auto-calculated
  - Single "Sizer" tab: inputs LEFT, auto-calculations RIGHT
  - Hidden "Colchis Data" sheet with leverage + pricing VLOOKUP tables
  - Zillow Market Data for ZHVI lookups

Cell references preserved for dealfit.py compatibility:
  C5=Deal Type, C6=Transaction, C7=Loan Term, C8=Deal Product
  C11=Address, C12=City, E12=State, C13=ZIP, C14=Property Type, C15=Units
  C16=Sqft, E16=Lot Size, C17=Year Built
  C20=Purchase Price, C21=Purchase Date, C22=As-Is Value
  C23=ARV, C24=Rehab Budget, C25=Total Project Cost
  C28=Initial Loan, C29=Rehab Holdback, C30=Interest Reserve
  C31=Total Loan
  C34=Entity, C35=# Guarantors
  C38=G1 Name, C39=G1 FICO
  C42=G2 Name, C43=G2 FICO
  C46=# Projects, C47=Similar Experience
  E20=ZHVI, E21=ZHVI Ratio

Output: assets/AS_Capital_Sizer.xlsx
"""

import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins

# ---------------------------------------------------------------------------
# PATH CONFIGURATION
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR = os.path.join(SCRIPT_DIR, "assets")
OUTPUT_PATH = os.path.join(ASSETS_DIR, "AS_Capital_Sizer.xlsx")
EXISTING_PATH = OUTPUT_PATH

# ---------------------------------------------------------------------------
# COLOUR PALETTE
# ---------------------------------------------------------------------------
DEEP_BLUE   = "0B5394"
TEAL        = "087496"
POWDER_BLUE = "A3D5E0"
LIGHT_BLUE  = "E0F0F8"
DARK_TEXT    = "2C3E50"
WHITE        = "FFFFFF"
PASS_GREEN   = "27AE60"
FAIL_RED     = "E74C3C"
LIGHT_GRAY   = "F2F2F2"
MED_GRAY     = "999999"
DARK_GRAY    = "444444"
BLACK        = "000000"
PASS_BG      = "C6EFCE"
PASS_FG      = "006100"
FAIL_BG      = "FFC7CE"
FAIL_FG      = "9C0006"

# ---------------------------------------------------------------------------
# REUSABLE STYLE OBJECTS
# ---------------------------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color=BLACK),
    right=Side(style="thin", color=BLACK),
    top=Side(style="thin", color=BLACK),
    bottom=Side(style="thin", color=BLACK),
)
NO_BORDER = Border()

FONT_TITLE         = Font(name="Calibri", size=16, bold=True, color=WHITE)
FONT_SECTION        = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_SUBSECTION     = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
FONT_LABEL          = Font(name="Calibri", size=10, color=DARK_TEXT)
FONT_INPUT          = Font(name="Calibri", size=10, color=BLACK)
FONT_COMPUTED       = Font(name="Calibri", size=10, color=DARK_TEXT)
FONT_COMPUTED_BOLD  = Font(name="Calibri", size=10, bold=True, color=DARK_TEXT)
FONT_BIG_RESULT     = Font(name="Calibri", size=12, bold=True, color=WHITE)
FONT_NOTE           = Font(name="Calibri", size=9, italic=True, color=MED_GRAY)
FONT_PASS           = Font(name="Calibri", size=10, bold=True, color=PASS_FG)
FONT_FAIL           = Font(name="Calibri", size=10, bold=True, color=FAIL_FG)
FONT_REF_HEADER     = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_REF_SECTION    = Font(name="Calibri", size=11, bold=True, color=DEEP_BLUE)
FONT_REF            = Font(name="Calibri", size=9, color=DARK_TEXT)
FONT_REF_BOLD       = Font(name="Calibri", size=9, bold=True, color=DARK_TEXT)

FILL_DEEP_BLUE  = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
FILL_TEAL       = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
FILL_POWDER     = PatternFill(start_color=POWDER_BLUE, end_color=POWDER_BLUE, fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_WHITE      = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_PASS       = PatternFill(start_color=PASS_BG, end_color=PASS_BG, fill_type="solid")
FILL_FAIL       = PatternFill(start_color=FAIL_BG, end_color=FAIL_BG, fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT  = Alignment(horizontal="right", vertical="center")

FMT_CURRENCY     = '$#,##0'
FMT_CURRENCY_DEC = '$#,##0.00'
FMT_PCT          = '0.0%'
FMT_RATE         = '0.000%'
FMT_INT          = '#,##0'
FMT_DATE         = 'MM/DD/YYYY'
FMT_TEXT         = '@'

# ---------------------------------------------------------------------------
# DROPDOWN LISTS
# ---------------------------------------------------------------------------
STATES = "AL,AK,AZ,AR,CA,CO,CT,DC,DE,FL,GA,HI,ID,IL,IN,IA,KS,KY,LA,ME,MD,MA,MI,MN,MS,MO,MT,NE,NV,NH,NJ,NM,NY,NC,ND,OH,OK,OR,PA,RI,SC,SD,TN,TX,UT,VT,VA,WA,WV,WI,WY"

DROPDOWNS = {
    "deal_type":     "Fix & Flip,Bridge,Fix & Hold,Ground Up Construction",
    "transaction":   "Purchase,Refinance (Rate & Term),Refinance (Cash Out)",
    "loan_term":     "12 Months,18 Months,24 Months",
    "deal_product":  "Light Rehab,Heavy Rehab,Bridge,Construction",
    "property_type": "SFR,Townhome,Condo,PUD,2-4 Unit,5-10 MFR,11-20 MFR,21-50 MFR",
    "state":         STATES,
    "experience":    "Yes,No,Limited",
    "guarantors":    "1,2,3,4",
    "yes_no":        "Yes,No",
}


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def _apply_input_style(cell, fmt=None):
    """White bg, thin border -- an input cell."""
    cell.fill = FILL_WHITE
    cell.border = THIN_BORDER
    cell.font = FONT_INPUT
    cell.alignment = ALIGN_LEFT
    if fmt:
        cell.number_format = fmt


def _apply_computed_style(cell, fmt=None, bold=False):
    """Light blue bg -- a computed/formula cell."""
    cell.fill = FILL_LIGHT_BLUE
    cell.border = THIN_BORDER
    cell.font = FONT_COMPUTED_BOLD if bold else FONT_COMPUTED
    cell.alignment = ALIGN_RIGHT
    if fmt:
        cell.number_format = fmt


def _apply_label_style(cell):
    cell.font = FONT_LABEL
    cell.alignment = ALIGN_LEFT


def _section_header(ws, row, col_start, col_end, text):
    """Teal section header spanning col_start:col_end."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = FONT_SECTION
    cell.fill = FILL_TEAL
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for c in range(col_start + 1, col_end + 1):
        mc = ws.cell(row=row, column=c)
        mc.fill = FILL_TEAL
        mc.border = THIN_BORDER


def _sub_header(ws, row, col_start, col_end, text):
    """Lighter sub-header."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = FONT_SUBSECTION
    cell.fill = FILL_LIGHT_GRAY
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for c in range(col_start + 1, col_end + 1):
        mc = ws.cell(row=row, column=c)
        mc.fill = FILL_LIGHT_GRAY
        mc.border = THIN_BORDER


def _title_row(ws, row, col_start, col_end, text):
    """Deep blue full-width title."""
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = FONT_TITLE
    cell.fill = FILL_DEEP_BLUE
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    for c in range(col_start + 1, col_end + 1):
        mc = ws.cell(row=row, column=c)
        mc.fill = FILL_DEEP_BLUE
        mc.border = THIN_BORDER


def _add_dropdown(ws, cell_range, formula_list, prompt_title="", prompt_body=""):
    dv = DataValidation(
        type="list",
        formula1=f'"{formula_list}"',
        allow_blank=True,
        showDropDown=False,
    )
    if prompt_title:
        dv.prompt = prompt_body
        dv.promptTitle = prompt_title
        dv.showInputMessage = True
    dv.add(cell_range)
    ws.add_data_validation(dv)


def _label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    _apply_label_style(c)
    return c


def _input_cell(ws, row, col, fmt=None):
    c = ws.cell(row=row, column=col)
    _apply_input_style(c, fmt)
    return c


def _formula_cell(ws, row, col, formula, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    _apply_computed_style(c, fmt, bold)
    return c


def _merged_input(ws, row, col_start, col_end, fmt=None):
    ws.merge_cells(
        start_row=row, start_column=col_start,
        end_row=row, end_column=col_end
    )
    c = ws.cell(row=row, column=col_start)
    _apply_input_style(c, fmt)
    for cc in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=cc).border = THIN_BORDER
    return c


def _result_cell(ws, row, col, formula, fmt=None):
    """Pass/Fail result cell."""
    c = ws.cell(row=row, column=col, value=formula)
    c.border = THIN_BORDER
    c.alignment = ALIGN_CENTER
    c.font = FONT_COMPUTED_BOLD
    if fmt:
        c.number_format = fmt
    return c


def _col_header(ws, row, col, text):
    """Column header with powder blue background."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_SUBSECTION
    c.fill = FILL_POWDER
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    return c


# ============================================================================
# SIZER SHEET  (Eastview architecture with Colchis data)
# ============================================================================
#
# LEFT SIDE  (cols B-E): All user inputs
# RIGHT SIDE (cols G-K): Auto-calculations, sizing, pricing, checks
#
# Eastview's key innovation: the Leverage lookup table is keyed by
#   "LoanPurpose / BorrowerClass"  e.g. "Purchase / A+"
# This avoids the FICO-bucket + experience-tier combinatorial explosion
# that caused v3's VLOOKUP failures.
#
# For Colchis, we adapt this by adding FICO as a third dimension:
#   "Product|LoanPurpose|Class|FICO"
# so e.g. "Construction|Refinance (Rate & Term)|A+|700-739"
#
# ============================================================================

def build_sizer_sheet(wb):
    ws = wb.active
    ws.title = "Sizer"

    # Column widths
    widths = {1: 3, 2: 24, 3: 20, 4: 20, 5: 20, 6: 3,
              7: 28, 8: 20, 9: 20, 10: 20, 11: 3}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 36

    # ---- Title ----
    _title_row(ws, 2, 1, 11, "A&S CAPITAL SIZER")

    # ==================================================================
    # LEFT: DEAL INFORMATION (rows 4-8)
    # ==================================================================
    _section_header(ws, 4, 2, 5, "DEAL INFORMATION")

    _label(ws, 5, 2, "Deal Type")
    _input_cell(ws, 5, 3)
    _add_dropdown(ws, "C5", DROPDOWNS["deal_type"], "Deal Type", "Select deal type")

    _label(ws, 6, 2, "Transaction Type")
    _input_cell(ws, 6, 3)
    _add_dropdown(ws, "C6", DROPDOWNS["transaction"], "Transaction", "Select transaction type")

    _label(ws, 7, 2, "Loan Term")
    _input_cell(ws, 7, 3)
    _add_dropdown(ws, "C7", DROPDOWNS["loan_term"], "Loan Term", "Select loan term")

    _label(ws, 8, 2, "Deal Product")
    _input_cell(ws, 8, 3)
    _add_dropdown(ws, "C8", DROPDOWNS["deal_product"], "Product", "Select deal product")

    # ==================================================================
    # RIGHT: BORROWER CLASSIFICATION (rows 4-8)
    # Mirrors Eastview's classification system
    # ==================================================================
    _section_header(ws, 4, 7, 10, "BORROWER CLASSIFICATION")

    # Qualifying Experience (like Eastview T38/T39)
    _label(ws, 5, 7, "Qualifying Experience")
    _formula_cell(ws, 5, 8, '=C46', FMT_INT)

    # Borrower Classification (like Eastview T40): A+/A/B/N/A
    # For Colchis Construction: 6+=A+, 4-5=A, 0-3=Ineligible
    # For all others: 8+=A+, 4-7=A, 0-3=B
    _label(ws, 6, 7, "Borrower Classification")
    _formula_cell(
        ws, 6, 8,
        '=IF(C46="","",IF(C8="Construction",'
        'IF(C46>=6,"A+",IF(C46>=4,"A","Ineligible")),'
        'IF(C46>=8,"A+",IF(C46>=4,"A",IF(C46>=1,"B","Ineligible")))))',
    )

    # FICO Bucket (bucketed for lookup)
    _label(ws, 7, 7, "FICO Bucket")
    _formula_cell(
        ws, 7, 8,
        '=IF(C39="","",IF(C39>=740,"740+",IF(C39>=700,"700-739",'
        'IF(C39>=680,"680-699","<680 (Ineligible)"))))',
    )

    # Leverage Lookup Key:  "Product|Transaction|Class|FICO"
    _label(ws, 8, 7, "Leverage Lookup Key")
    _formula_cell(
        ws, 8, 8,
        '=IF(OR(C8="",C6="",H6="",H7=""),"",C8&"|"&C6&"|"&H6&"|"&H7)',
    )

    # ==================================================================
    # LEFT: PROPERTY INFORMATION (rows 10-17)
    # ==================================================================
    _section_header(ws, 10, 2, 5, "PROPERTY INFORMATION")

    _label(ws, 11, 2, "Property Address")
    _merged_input(ws, 11, 3, 4)

    _label(ws, 12, 2, "City")
    _input_cell(ws, 12, 3)
    _label(ws, 12, 4, "State")
    _input_cell(ws, 12, 5)
    _add_dropdown(ws, "E12", DROPDOWNS["state"], "State", "Select state")

    _label(ws, 13, 2, "ZIP Code")
    _input_cell(ws, 13, 3, FMT_INT)   # Numeric format so VLOOKUP works

    _label(ws, 14, 2, "Property Type")
    _input_cell(ws, 14, 3)
    _add_dropdown(ws, "C14", DROPDOWNS["property_type"], "Property Type", "Select property type")

    _label(ws, 15, 2, "# Units")
    _input_cell(ws, 15, 3, FMT_INT)

    _label(ws, 16, 2, "Square Footage")
    _input_cell(ws, 16, 3, FMT_INT)
    _label(ws, 16, 4, "Lot Size (SF)")
    _input_cell(ws, 16, 5, FMT_INT)

    _label(ws, 17, 2, "Year Built")
    _formula_cell(
        ws, 17, 3,
        '=IF(C5="Ground Up Construction",YEAR(TODAY()),"")',
        "0"
    )
    note_yb = ws.cell(row=17, column=4, value="Auto-fills for GUC")
    note_yb.font = FONT_NOTE
    note_yb.alignment = ALIGN_LEFT

    # ==================================================================
    # RIGHT: LEVERAGE LIMITS (rows 10-19)
    # VLOOKUP from Colchis Data using the classification-based key
    # ==================================================================
    _section_header(ws, 10, 7, 10, "COLCHIS LEVERAGE LIMITS")

    _col_header(ws, 11, 8, "Max %")
    _col_header(ws, 11, 9, "Max $ Amount")

    # Max LTV (As-Is)
    _label(ws, 12, 7, "Max LTV (As-Is)")
    _formula_cell(
        ws, 12, 8,
        '=IFERROR(VLOOKUP(H8,\'Colchis Data\'!A:E,2,FALSE),"")',
        FMT_PCT
    )
    _formula_cell(ws, 12, 9, '=IFERROR(H12*C22,"")', FMT_CURRENCY)

    # Max LTC
    _label(ws, 13, 7, "Max LTC")
    _formula_cell(
        ws, 13, 8,
        '=IFERROR(VLOOKUP(H8,\'Colchis Data\'!A:E,3,FALSE),"")',
        FMT_PCT
    )
    _formula_cell(ws, 13, 9, '=IFERROR(H13*C25,"")', FMT_CURRENCY)

    # Max LTARV
    _label(ws, 14, 7, "Max LTARV")
    _formula_cell(
        ws, 14, 8,
        '=IFERROR(VLOOKUP(H8,\'Colchis Data\'!A:E,4,FALSE),"")',
        FMT_PCT
    )
    _formula_cell(ws, 14, 9, '=IFERROR(H14*C23,"")', FMT_CURRENCY)

    # Max TLTC (Total Loan-to-Cost, for Purchase only -- like Eastview col G/K)
    _label(ws, 15, 7, "Max TLTC")
    _formula_cell(
        ws, 15, 8,
        '=IFERROR(VLOOKUP(H8,\'Colchis Data\'!A:E,5,FALSE),"")',
        FMT_PCT
    )
    _formula_cell(ws, 15, 9, '=IFERROR(H15*C25,"")', FMT_CURRENCY)

    # Interest Reserves auto-calc (5% of LTARV max, like Eastview T26)
    _label(ws, 16, 7, "Interest Reserves (5%)")
    _formula_cell(
        ws, 16, 9,
        '=IFERROR(IF(H14="","",(H14*C23)*0.05),"")',
        FMT_CURRENCY
    )
    note_ir = ws.cell(row=16, column=10, value="5% of (LTARV × ARV)")
    note_ir.font = FONT_NOTE

    # --- CORE SIZING (like Eastview K81) ---
    # Eastview formula for Initial Loan Amount:
    #   Purchase: MIN(AsIs*LTV, LTC*PurchasePrice, ARV*LTARV - Rehab - IR, TLTC*TotalCost - Rehab - IR)
    #   Refi:     MIN(AsIs*LTV, ARV*LTARV - Rehab - IR)
    # For Construction (Colchis): LTV is irrelevant, use LTC/LTARV only
    #   Purchase: MIN(LTC*TotalCost, ARV*LTARV) - Rehab - IR  (front end = LTC, back end = LTARV)
    #   Refi:     MIN(AsIs*LTV, ARV*LTARV - Rehab - IR)

    _label(ws, 18, 7, "Max Initial Loan Amount")
    ws.cell(row=18, column=7).font = FONT_COMPUTED_BOLD
    # This is the critical formula. It replicates Eastview's K81 logic:
    _formula_cell(
        ws, 18, 9,
        '=IFERROR(ROUNDDOWN(MIN('
        # Front End constraint
        'IF(C8="Construction",'
          # Construction: no LTV front-end for refi; for purchase use LTC×TotalCost
          'IF(C6="Purchase",H13*C25,9999999999),'
          # Others: AsIs × LTV (and for Purchase also LTC × PurchasePrice)
          'IF(C6="Purchase",MIN(C22*H12,H13*C20),C22*H12)'
        '),'
        # Back End constraint
        'IF(C8="Construction",'
          # Construction: LTARV×ARV only (no LTV here!) minus rehab minus IR
          '(H14*C23)-C24-I16,'
          # Others: (LTARV×ARV minus rehab minus IR) and for Purchase also (TLTC×TotalCost - rehab - IR)
          'IF(C6="Purchase",'
            'MIN(H14*C23,H15*C25)-C24-I16,'
            '(H14*C23)-C24-I16'
          ')'
        ')'
        '),0),"")',
        FMT_CURRENCY, bold=True
    )

    # Max Loan Cap
    _label(ws, 19, 7, "Max Loan Amount Cap")
    cap_cell = ws.cell(row=19, column=9, value=3500000)
    cap_cell.number_format = FMT_CURRENCY
    cap_cell.font = FONT_COMPUTED
    cap_cell.fill = FILL_LIGHT_BLUE
    cap_cell.border = THIN_BORDER
    cap_cell.alignment = ALIGN_RIGHT

    # ==================================================================
    # LEFT: VALUATION (rows 19-25) with ZHVI on D-E
    # ==================================================================
    _section_header(ws, 19, 2, 5, "VALUATION")

    _label(ws, 20, 2, "Purchase Price")
    _input_cell(ws, 20, 3, FMT_CURRENCY)
    _label(ws, 20, 4, "ZHVI (Zillow)")
    _formula_cell(
        ws, 20, 5,
        '=IFERROR(VLOOKUP(C13,\'Zillow Market Data\'!C:AN,38,FALSE),"")',
        FMT_CURRENCY
    )

    _label(ws, 21, 2, "Purchase Date")
    _input_cell(ws, 21, 3, FMT_DATE)
    _label(ws, 21, 4, "Value / ZHVI Ratio")
    _formula_cell(ws, 21, 5, '=IFERROR(C22/E20,"")', '0.00"x"')

    _label(ws, 22, 2, "As-Is Value")
    _input_cell(ws, 22, 3, FMT_CURRENCY)
    _label(ws, 22, 4, "Deal vs Market")
    _formula_cell(
        ws, 22, 5,
        '=IF(E21="","",IF(E21>3,"HIGH RISK",IF(E21>2,"ELEVATED","NORMAL")))'
    )

    _label(ws, 23, 2, "After Repair Value (ARV)")
    _input_cell(ws, 23, 3, FMT_CURRENCY)

    _label(ws, 24, 2, "Rehab Budget")
    _input_cell(ws, 24, 3, FMT_CURRENCY)

    _label(ws, 25, 2, "Total Project Cost")
    _formula_cell(ws, 25, 3, "=C20+C24", FMT_CURRENCY, bold=True)

    # ==================================================================
    # RIGHT: LOAN PROCEEDS (rows 21-30)
    # Mirrors Eastview's Loan Proceeds section
    # ==================================================================
    _section_header(ws, 21, 7, 10, "LOAN PROCEEDS")

    _col_header(ws, 22, 8, "Borrower Req.")
    _col_header(ws, 22, 9, "Guidelines Max")

    # Initial Loan Amount
    _label(ws, 23, 7, "Initial Loan Amount")
    _formula_cell(ws, 23, 8, '=C28', FMT_CURRENCY)
    # Guidelines max = MIN(max initial loan, max loan cap)
    _formula_cell(ws, 23, 9,
        '=IFERROR(IF(I18="","",MIN(MAX(I18,0),I19)),"")',
        FMT_CURRENCY)

    # Financed Rehab Budget
    _label(ws, 24, 7, "Financed Rehab Budget")
    _formula_cell(ws, 24, 8, '=C29', FMT_CURRENCY)
    # Max rehab = full rehab budget if initial loan > 0
    _formula_cell(ws, 24, 9,
        '=IFERROR(IF(I23=0,0,ROUNDDOWN(C24,0)),"")',
        FMT_CURRENCY)

    # Interest Reserve
    _label(ws, 25, 7, "Interest Reserve")
    _formula_cell(ws, 25, 8, '=C30', FMT_CURRENCY)
    # Max IR from auto-calc
    _formula_cell(ws, 25, 9, '=IFERROR(IF(I23=0,0,I16),"")', FMT_CURRENCY)

    # Total Loan Amount (bold)
    _label(ws, 26, 7, "Total Loan Amount")
    ws.cell(row=26, column=7).font = FONT_COMPUTED_BOLD
    _formula_cell(ws, 26, 8, '=SUM(H23:H25)', FMT_CURRENCY, bold=True)
    _formula_cell(ws, 26, 9, '=IFERROR(IF(I23=0,0,I23+I24+I25),"")', FMT_CURRENCY, bold=True)

    # Actual ratios
    _label(ws, 28, 7, "Actual LTV")
    _formula_cell(ws, 28, 8, '=IFERROR(H23/C22,"")', FMT_PCT)
    _formula_cell(ws, 28, 9, '=IFERROR(I23/C22,"")', FMT_PCT)

    _label(ws, 29, 7, "Actual LTC (Total)")
    _formula_cell(ws, 29, 8, '=IFERROR(H26/C25,"")', FMT_PCT)
    _formula_cell(ws, 29, 9, '=IFERROR(I26/C25,"")', FMT_PCT)

    _label(ws, 30, 7, "Actual LTARV")
    _formula_cell(ws, 30, 8, '=IFERROR(H26/C23,"")', FMT_PCT)
    _formula_cell(ws, 30, 9, '=IFERROR(I26/C23,"")', FMT_PCT)

    # ==================================================================
    # LEFT: LOAN REQUEST (rows 27-31)
    # ==================================================================
    _section_header(ws, 27, 2, 5, "LOAN REQUEST")

    _label(ws, 28, 2, "Initial Loan Amount")
    _input_cell(ws, 28, 3, FMT_CURRENCY)

    _label(ws, 29, 2, "Rehab Holdback")
    _input_cell(ws, 29, 3, FMT_CURRENCY)

    _label(ws, 30, 2, "Interest Reserve")
    _input_cell(ws, 30, 3, FMT_CURRENCY)

    _label(ws, 31, 2, "Total Loan Amount")
    _formula_cell(ws, 31, 3, "=C28+C29+C30", FMT_CURRENCY, bold=True)

    # Leverage ratios on right of loan request
    _label(ws, 28, 4, "LTV")
    _formula_cell(ws, 28, 5, '=IFERROR(C28/C22,"")', FMT_PCT)

    _label(ws, 29, 4, "LTC")
    _formula_cell(ws, 29, 5, '=IFERROR(C31/C25,"")', FMT_PCT)

    _label(ws, 30, 4, "LTARV")
    _formula_cell(ws, 30, 5, '=IFERROR(C31/C23,"")', FMT_PCT)

    # ==================================================================
    # RIGHT: COLCHIS PRICING (rows 32-44)
    # ==================================================================
    _section_header(ws, 32, 7, 10, "COLCHIS PRICING")

    # LTC bucket for pricing
    _label(ws, 33, 7, "Actual LTC %")
    _formula_cell(ws, 33, 8, '=IFERROR(H26/C25,"")', FMT_PCT)

    _label(ws, 34, 7, "LTC Bucket")
    _formula_cell(
        ws, 34, 8,
        '=IF(H33="","",IF(H33<=0.7,"<=70.0%",IF(H33<=0.75,"<=75.0%",'
        'IF(H33<=0.8,"<=80.0%",IF(H33<=0.85,"<=85.0%",'
        'IF(H33<=0.9,"<=90.0%","<=95.0%"))))))',
    )

    # Pricing Key: "Product|FICO|LTCBucket"
    _label(ws, 35, 7, "Pricing Key")
    _formula_cell(
        ws, 35, 8,
        '=IF(OR(C8="",H7="",H34=""),"",C8&"|"&H7&"|"&H34)',
    )

    # Base Rate
    _label(ws, 36, 7, "Base Rate")
    _formula_cell(
        ws, 36, 8,
        '=IFERROR(VLOOKUP(H35,\'Colchis Data\'!G:H,2,FALSE),"")',
        FMT_RATE
    )

    # Adjustments
    _label(ws, 37, 7, "Experience Adj.")
    _formula_cell(
        ws, 37, 8,
        '=IF(H6="","",'
        'IF(H6="A+",-0.0025,'
        'IF(H6="B",0.0025,0)))',
        FMT_RATE
    )
    ws.cell(row=37, column=9, value="A+: -25bps / B: +25bps").font = FONT_NOTE

    _label(ws, 38, 7, "Term Adj.")
    _formula_cell(
        ws, 38, 8,
        '=IF(C7="","",IF(C7="24 Months",0.00125,0))',
        FMT_RATE
    )
    ws.cell(row=38, column=9, value="24mo: +12.5bps").font = FONT_NOTE

    _label(ws, 39, 7, "Transaction Adj.")
    _formula_cell(
        ws, 39, 8,
        '=IF(C6="","",IF(C6="Refinance (Cash Out)",0.0025,0))',
        FMT_RATE
    )
    ws.cell(row=39, column=9, value="Cash-out refi: +25bps").font = FONT_NOTE

    _label(ws, 40, 7, "Loan Size Adj.")
    _formula_cell(
        ws, 40, 8,
        '=IF(H26="","",IF(H26>3000000,0.00125,0))',
        FMT_RATE
    )
    ws.cell(row=40, column=9, value=">$3M: +12.5bps").font = FONT_NOTE

    _label(ws, 41, 7, "State / ZHVI Adj.")
    _formula_cell(
        ws, 41, 8,
        '=IFERROR('
        'IF(E12="","",IF(OR(E12="NY",E12="NJ",E12="CT"),0.0025,IF(E12="CA",-0.00125,0)))'
        '+IF(AND(E21<>"",ISNUMBER(E21)),IF(E21>3,0.00375,IF(E21>2,0.00125,0)),0)'
        ',"")',
        FMT_RATE
    )
    ws.cell(row=41, column=9, value="NY/NJ/CT +25bp; CA -12.5bp; ZHVI").font = FONT_NOTE

    # ALL-IN BUY RATE
    _label(ws, 43, 7, "ALL-IN BUY RATE")
    ws.cell(row=43, column=7).font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
    c_rate = ws.cell(row=43, column=8, value='=IFERROR(H36+H37+H38+H39+H40+H41,"")')
    c_rate.font = FONT_BIG_RESULT
    c_rate.fill = FILL_DEEP_BLUE
    c_rate.number_format = FMT_RATE
    c_rate.alignment = ALIGN_CENTER
    c_rate.border = THIN_BORDER

    _label(ws, 44, 7, "A&S Sell Rate (+50bps)")
    _formula_cell(ws, 44, 8, '=IFERROR(H43+0.005,"")', FMT_RATE, bold=True)

    # ==================================================================
    # LEFT: BORROWER INFORMATION (rows 33-47)
    # ==================================================================
    _section_header(ws, 33, 2, 5, "BORROWER INFORMATION")

    _label(ws, 34, 2, "Borrowing Entity")
    _merged_input(ws, 34, 3, 4)

    _label(ws, 35, 2, "# Guarantors")
    _input_cell(ws, 35, 3)
    _add_dropdown(ws, "C35", DROPDOWNS["guarantors"], "Guarantors", "Number of guarantors")

    # Guarantor 1
    _sub_header(ws, 37, 2, 3, "GUARANTOR 1")
    _label(ws, 38, 2, "Full Name")
    _merged_input(ws, 38, 3, 4)
    _label(ws, 39, 2, "FICO Score")
    _input_cell(ws, 39, 3, FMT_INT)

    # Guarantor 2
    _sub_header(ws, 41, 2, 3, "GUARANTOR 2")
    _label(ws, 42, 2, "Full Name")
    _merged_input(ws, 42, 3, 4)
    _label(ws, 43, 2, "FICO Score")
    _input_cell(ws, 43, 3, FMT_INT)

    # ==================================================================
    # EXPERIENCE (rows 45-47)
    # ==================================================================
    _section_header(ws, 45, 2, 5, "EXPERIENCE")

    _label(ws, 46, 2, "# Completed Projects")
    _input_cell(ws, 46, 3, FMT_INT)

    _label(ws, 47, 2, "Similar Experience")
    _input_cell(ws, 47, 3)
    _add_dropdown(ws, "C47", DROPDOWNS["experience"], "Experience", "Similar project experience?")

    # ==================================================================
    # RIGHT: GUIDELINE CHECK (rows 46-58)
    # ==================================================================
    _section_header(ws, 46, 7, 10, "GUIDELINE CHECK")

    _col_header(ws, 47, 8, "Value")
    _col_header(ws, 47, 9, "Result")

    # Min FICO (680)
    _label(ws, 48, 7, "Min FICO (680)")
    _formula_cell(ws, 48, 8, '=C39', FMT_INT)
    _result_cell(ws, 48, 9, '=IF(C39="","",IF(C39>=680,"PASS","FAIL"))')

    # Max Loan ($3.5M)
    _label(ws, 49, 7, "Max Loan ($3.5M)")
    _formula_cell(ws, 49, 8, '=I26', FMT_CURRENCY)
    _result_cell(ws, 49, 9, '=IF(I26="","",IF(I26<=3500000,"PASS","FAIL"))')

    # Min Loan ($100K)
    _label(ws, 50, 7, "Min Loan ($100K)")
    _formula_cell(ws, 50, 8, '=I26', FMT_CURRENCY)
    _result_cell(ws, 50, 9, '=IF(I26="","",IF(I26>=100000,"PASS","FAIL"))')

    # State Eligible
    _label(ws, 51, 7, "State Eligible")
    _formula_cell(ws, 51, 8, '=E12')
    _result_cell(ws, 51, 9, '=IF(E12="","",IF(E12="IL","FAIL","PASS"))')

    # Borrower Eligible (classification check)
    _label(ws, 52, 7, "Borrower Eligible")
    _formula_cell(ws, 52, 8, '=H6')
    _result_cell(ws, 52, 9,
        '=IF(H6="","",IF(H6="Ineligible","FAIL",'
        'IF(H12="","FAIL","PASS")))')

    # Max Initial Loan check (like Eastview T80)
    _label(ws, 53, 7, "Max Initial Loan (Sizer)")
    _formula_cell(ws, 53, 8, '=H23', FMT_CURRENCY)
    _result_cell(ws, 53, 9, '=IF(OR(H23="",I23=""),"",IF(H23<=I23,"PASS","FAIL"))')

    # Max Total Loan check (like Eastview T82)
    _label(ws, 54, 7, "Max Total Loan (Sizer)")
    _formula_cell(ws, 54, 8, '=H26', FMT_CURRENCY)
    _result_cell(ws, 54, 9, '=IF(OR(H26="",I26=""),"",IF(H26<=I26,"PASS","FAIL"))')

    # Total LTC Cap (90%, like Eastview W90)
    _label(ws, 55, 7, "Total LTC Cap (90%)")
    _formula_cell(ws, 55, 8, '=IFERROR(H26/C25,"")', FMT_PCT)
    _result_cell(ws, 55, 9, '=IF(H55="","",IF(H55<=0.9,"PASS","FAIL"))')

    # ZHVI Check
    _label(ws, 56, 7, "ZHVI > 300% (High Risk)")
    _formula_cell(ws, 56, 8, '=E21')
    _result_cell(ws, 56, 9, '=IF(E21="","",IF(E21>3,"WARN","PASS"))')

    # MASTER CHECK
    _label(ws, 58, 7, "MASTER CHECK")
    ws.cell(row=58, column=7).font = Font(name="Calibri", size=11, bold=True, color=DARK_TEXT)
    c_master = ws.cell(
        row=58, column=9,
        value='=IF(COUNTBLANK(I48:I56)=9,"",IF(COUNTIF(I48:I56,"FAIL")>0,"FAIL","PASS"))'
    )
    c_master.border = THIN_BORDER
    c_master.alignment = ALIGN_CENTER
    c_master.font = Font(name="Calibri", size=12, bold=True, color=DARK_TEXT)

    # ---- Note row ----
    ws.merge_cells("B50:E50")
    note = ws.cell(
        row=50, column=2,
        value="Complete all fields. Auto-sizing results appear on the right."
    )
    note.font = FONT_NOTE
    note.alignment = ALIGN_CENTER

    # ---- Print setup ----
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_area = "A1:K60"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return ws


# ============================================================================
# COLCHIS DATA (HIDDEN lookup table)
# ============================================================================

def build_colchis_data_sheet(wb):
    """
    Hidden sheet with two lookup tables:
      Columns A-E: Leverage lookup
        Key = "Product|Transaction|Class|FICO"
        Col B = Max LTV, Col C = Max LTC, Col D = Max LTARV, Col E = Max TLTC
      Columns G-H: Pricing lookup
        Key = "Product|FICO|LTCBucket"
        Col H = Base Rate
    """
    ws = wb.create_sheet("Colchis Data")

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["H"].width = 12

    # ---- LEVERAGE TABLE HEADER ----
    for col_idx, hdr in enumerate(["Lookup Key", "Max LTV", "Max LTC", "Max LTARV", "Max TLTC"], 1):
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font = FONT_REF_BOLD
        c.fill = FILL_POWDER
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER

    # ---- PRICING TABLE HEADER ----
    for col_idx, hdr in [(7, "Pricing Key"), (8, "Base Rate")]:
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font = FONT_REF_BOLD
        c.fill = FILL_POWDER
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER

    # ================================================================
    # LEVERAGE DATA
    # Key: "Product|Transaction|BorrowerClass|FICOBucket"
    # Values: MaxLTV, MaxLTC, MaxLTARV, MaxTLTC
    #
    # Colchis guidelines by product:
    #   Construction: experience 6+=A+, 4-5=A, 0-3=Ineligible
    #   Heavy Rehab: experience 8+=A+, 4-7=A, 0-3=Ineligible
    #   Light Rehab: experience 8+=A+, 4-7=A, 0-3=B
    #   Bridge: experience 8+=A+, 4-7=A, 0-3=B
    #
    # For each Product×Class×FICO, we provide leverage for all 3 transaction types.
    # For Refi, LTC and TLTC are N/A (set to 0 so they don't constrain).
    # For Bridge, LTC/LTARV/TLTC are always 0 (LTV only).
    # ================================================================

    leverage_data = []

    # Helper to generate Purchase + R&T + C/O rows for a given product/class/fico
    def _add_rows(product, cls, fico, ltv, ltc, ltarv, tltc=None):
        """Add leverage rows for all 3 transaction types."""
        if tltc is None:
            tltc = ltc  # Default TLTC = LTC for Purchase
        # Purchase: all constraints active
        leverage_data.append((f"{product}|Purchase|{cls}|{fico}", ltv, ltc, ltarv, tltc))
        # Refi R&T: only LTV and LTARV, no LTC/TLTC
        leverage_data.append((f"{product}|Refinance (Rate & Term)|{cls}|{fico}", ltv, 0.0, ltarv, 0.0))
        # Refi C/O: reduced LTV and LTARV (per Colchis: typically -5% for cash-out)
        co_ltv = max(ltv - 0.05, 0.0)
        co_ltarv = max(ltarv - 0.05, 0.0)
        leverage_data.append((f"{product}|Refinance (Cash Out)|{cls}|{fico}", co_ltv, 0.0, co_ltarv, 0.0))

    # ---- LIGHT REHAB ----
    # A+ (8+ exp)
    _add_rows("Light Rehab", "A+", "740+",    0.900, 0.925, 0.750, 0.925)
    _add_rows("Light Rehab", "A+", "700-739", 0.900, 0.925, 0.750, 0.925)
    _add_rows("Light Rehab", "A+", "680-699", 0.875, 0.900, 0.750, 0.900)
    # A (4-7 exp)
    _add_rows("Light Rehab", "A", "740+",    0.900, 0.925, 0.750, 0.925)
    _add_rows("Light Rehab", "A", "700-739", 0.900, 0.925, 0.750, 0.925)
    _add_rows("Light Rehab", "A", "680-699", 0.850, 0.875, 0.750, 0.875)
    # B (0-3 exp)
    _add_rows("Light Rehab", "B", "740+",    0.900, 0.900, 0.750, 0.900)
    _add_rows("Light Rehab", "B", "700-739", 0.875, 0.900, 0.750, 0.900)
    _add_rows("Light Rehab", "B", "680-699", 0.850, 0.850, 0.700, 0.850)

    # ---- HEAVY REHAB (0-3 = Ineligible, handled by classification formula) ----
    # A+
    _add_rows("Heavy Rehab", "A+", "740+",    0.800, 0.850, 0.700, 0.850)
    _add_rows("Heavy Rehab", "A+", "700-739", 0.800, 0.850, 0.700, 0.850)
    _add_rows("Heavy Rehab", "A+", "680-699", 0.750, 0.825, 0.650, 0.825)
    # A
    _add_rows("Heavy Rehab", "A", "740+",    0.800, 0.850, 0.700, 0.850)
    _add_rows("Heavy Rehab", "A", "700-739", 0.800, 0.850, 0.700, 0.850)
    _add_rows("Heavy Rehab", "A", "680-699", 0.750, 0.800, 0.650, 0.800)

    # ---- BRIDGE (LTV only -- LTC/LTARV/TLTC = 0) ----
    # A+
    leverage_data.append(("Bridge|Purchase|A+|740+", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|A+|740+", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|A+|740+", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Purchase|A+|700-739", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|A+|700-739", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|A+|700-739", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Purchase|A+|680-699", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|A+|680-699", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|A+|680-699", 0.650, 0.0, 0.0, 0.0))
    # A
    leverage_data.append(("Bridge|Purchase|A|740+", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|A|740+", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|A|740+", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Purchase|A|700-739", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|A|700-739", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|A|700-739", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Purchase|A|680-699", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|A|680-699", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|A|680-699", 0.650, 0.0, 0.0, 0.0))
    # B
    leverage_data.append(("Bridge|Purchase|B|740+", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|B|740+", 0.750, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|B|740+", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Purchase|B|700-739", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|B|700-739", 0.700, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|B|700-739", 0.650, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Purchase|B|680-699", 0.650, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Rate & Term)|B|680-699", 0.650, 0.0, 0.0, 0.0))
    leverage_data.append(("Bridge|Refinance (Cash Out)|B|680-699", 0.600, 0.0, 0.0, 0.0))

    # ---- CONSTRUCTION (0-3 = Ineligible) ----
    # For Construction, LTV is very low (60% on raw land). The real constraints are LTC and LTARV.
    # A+ (6+ exp)
    _add_rows("Construction", "A+", "740+",    0.600, 0.900, 0.700, 0.900)
    _add_rows("Construction", "A+", "700-739", 0.600, 0.900, 0.700, 0.900)
    _add_rows("Construction", "A+", "680-699", 0.600, 0.850, 0.700, 0.850)
    # A (4-5 exp)
    _add_rows("Construction", "A", "740+",    0.600, 0.850, 0.700, 0.850)
    _add_rows("Construction", "A", "700-739", 0.600, 0.850, 0.700, 0.850)
    _add_rows("Construction", "A", "680-699", 0.600, 0.825, 0.650, 0.825)

    # Write leverage data
    row = 2
    for key, ltv, ltc, ltarv, tltc in leverage_data:
        ws.cell(row=row, column=1, value=key).font = FONT_REF
        ws.cell(row=row, column=1).border = THIN_BORDER

        for ci, val in enumerate([ltv, ltc, ltarv, tltc], 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.number_format = FMT_PCT
            c.font = FONT_REF
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER

        row += 1

    # ================================================================
    # PRICING DATA (from Colchis RTL Pricing 2026-01-27)
    # Key format: "Product|FICO Bucket|LTC Bucket"
    # These are BASE rates before adjustments
    # ================================================================
    pricing_data = [
        # ---- Bridge ----
        ("Bridge|740+|<=70.0%",     0.07750),
        ("Bridge|740+|<=75.0%",     0.07750),
        ("Bridge|700-739|<=70.0%",  0.07750),
        ("Bridge|700-739|<=75.0%",  0.07750),
        ("Bridge|680-699|<=70.0%",  0.07875),

        # ---- Light Rehab ----
        ("Light Rehab|740+|<=70.0%",     0.07750),
        ("Light Rehab|740+|<=75.0%",     0.07750),
        ("Light Rehab|740+|<=80.0%",     0.07750),
        ("Light Rehab|740+|<=85.0%",     0.07875),
        ("Light Rehab|740+|<=90.0%",     0.08000),
        ("Light Rehab|740+|<=95.0%",     0.08250),
        ("Light Rehab|700-739|<=70.0%",  0.07750),
        ("Light Rehab|700-739|<=75.0%",  0.07750),
        ("Light Rehab|700-739|<=80.0%",  0.07875),
        ("Light Rehab|700-739|<=85.0%",  0.08000),
        ("Light Rehab|700-739|<=90.0%",  0.08125),
        ("Light Rehab|700-739|<=95.0%",  0.08375),
        ("Light Rehab|680-699|<=70.0%",  0.07875),
        ("Light Rehab|680-699|<=75.0%",  0.08000),
        ("Light Rehab|680-699|<=80.0%",  0.08125),
        ("Light Rehab|680-699|<=85.0%",  0.08250),
        ("Light Rehab|680-699|<=90.0%",  0.08375),

        # ---- Heavy Rehab ----
        ("Heavy Rehab|740+|<=70.0%",     0.08375),
        ("Heavy Rehab|740+|<=75.0%",     0.08375),
        ("Heavy Rehab|740+|<=80.0%",     0.08500),
        ("Heavy Rehab|740+|<=85.0%",     0.08625),
        ("Heavy Rehab|700-739|<=70.0%",  0.08375),
        ("Heavy Rehab|700-739|<=75.0%",  0.08500),
        ("Heavy Rehab|700-739|<=80.0%",  0.08625),
        ("Heavy Rehab|700-739|<=85.0%",  0.08750),
        ("Heavy Rehab|680-699|<=70.0%",  0.08625),
        ("Heavy Rehab|680-699|<=75.0%",  0.08750),
        ("Heavy Rehab|680-699|<=80.0%",  0.08875),
        ("Heavy Rehab|680-699|<=85.0%",  0.09000),

        # ---- Construction ----
        ("Construction|740+|<=70.0%",     0.08375),
        ("Construction|740+|<=75.0%",     0.08375),
        ("Construction|740+|<=80.0%",     0.08500),
        ("Construction|740+|<=85.0%",     0.08625),
        ("Construction|740+|<=90.0%",     0.08875),
        ("Construction|700-739|<=70.0%",  0.08375),
        ("Construction|700-739|<=75.0%",  0.08500),
        ("Construction|700-739|<=80.0%",  0.08625),
        ("Construction|700-739|<=85.0%",  0.08750),
        ("Construction|700-739|<=90.0%",  0.09000),
        ("Construction|680-699|<=70.0%",  0.08625),
        ("Construction|680-699|<=75.0%",  0.08750),
        ("Construction|680-699|<=80.0%",  0.08875),
        ("Construction|680-699|<=85.0%",  0.09000),
    ]

    pr_row = 2
    for key, rate in pricing_data:
        ws.cell(row=pr_row, column=7, value=key).font = FONT_REF
        ws.cell(row=pr_row, column=7).border = THIN_BORDER

        c_rate = ws.cell(row=pr_row, column=8, value=rate)
        c_rate.number_format = FMT_RATE
        c_rate.font = FONT_REF
        c_rate.border = THIN_BORDER
        c_rate.alignment = ALIGN_CENTER

        pr_row += 1

    # HIDE this sheet
    ws.sheet_state = 'hidden'
    return ws


# ============================================================================
# COLCHIS LEVERAGE (Human-Readable Reference)
# ============================================================================

def build_colchis_leverage_sheet(wb):
    ws = wb.create_sheet("Colchis Leverage")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 22

    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value="COLCHIS CAPITAL \u2014 LEVERAGE GUIDELINES")
    c.font = FONT_REF_HEADER
    c.fill = FILL_DEEP_BLUE
    c.alignment = ALIGN_CENTER
    for cc in range(2, 10):
        ws.cell(row=1, column=cc).fill = FILL_DEEP_BLUE

    def _grid(start_row, title, col_headers, data_rows, note=""):
        r = start_row
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        c = ws.cell(row=r, column=2, value=title)
        c.font = FONT_REF_SECTION
        c.fill = FILL_LIGHT_GRAY
        c.alignment = ALIGN_LEFT
        for cc in range(3, 6):
            ws.cell(row=r, column=cc).fill = FILL_LIGHT_GRAY
        r += 1

        if note:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            ws.cell(row=r, column=2, value=note).font = FONT_NOTE
            r += 1

        for i, h in enumerate(col_headers):
            c = ws.cell(row=r, column=2 + i, value=h)
            c.font = FONT_REF_BOLD
            c.fill = FILL_POWDER
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER
        r += 1

        for row_data in data_rows:
            for i, val in enumerate(row_data):
                c = ws.cell(row=r, column=2 + i, value=val)
                c.font = FONT_REF
                c.border = THIN_BORDER
                c.alignment = ALIGN_CENTER if i > 0 else ALIGN_LEFT
            r += 1
        return r + 1

    row = 3
    row = _grid(row,
        "SF \u2014 LIGHT REHAB (Purchase)",
        ["FICO", "A+ (8+ exp)", "A (4-7 exp)", "B (0-3 exp)"],
        [
            ["740+",    "90% / 92.5% / 75%", "90% / 92.5% / 75%", "90% / 90% / 75%"],
            ["700-739", "90% / 92.5% / 75%", "90% / 92.5% / 75%", "87.5% / 90% / 75%"],
            ["680-699", "87.5% / 90% / 75%", "85% / 87.5% / 75%", "85% / 85% / 70%"],
        ],
        note="Format: Max LTV / Max LTC / Max LTARV"
    )

    row = _grid(row,
        "SF \u2014 HEAVY REHAB",
        ["FICO", "A+ (8+ exp)", "A (4-7 exp)", "B (0-3 exp)"],
        [
            ["740+",    "80% / 85% / 70%",   "80% / 85% / 70%",   "N/A"],
            ["700-739", "80% / 85% / 70%",   "80% / 85% / 70%",   "N/A"],
            ["680-699", "75% / 82.5% / 65%", "75% / 80% / 65%",   "N/A"],
        ],
        note="Format: Max LTV / Max LTC / Max LTARV"
    )

    row = _grid(row,
        "SF \u2014 BRIDGE (No Rehab)",
        ["FICO", "A+ (8+ exp)", "A (4-7 exp)", "B (0-3 exp)"],
        [
            ["740+",    "75%", "75%", "75%"],
            ["700-739", "75%", "75%", "70%"],
            ["680-699", "70%", "70%", "65%"],
        ],
        note="Max LTV only"
    )

    row = _grid(row,
        "SF \u2014 CONSTRUCTION",
        ["FICO", "A+ (6+ exp)", "A (4-5 exp)", "B (0-3 exp)"],
        [
            ["740+",    "60% LTV / 90%* LTC / 70% LTARV", "60% LTV / 85% LTC / 70% LTARV", "N/A"],
            ["700-739", "60% LTV / 90%* LTC / 70% LTARV", "60% LTV / 85% LTC / 70% LTARV", "N/A"],
            ["680-699", "60% LTV / 85% LTC / 70% LTARV", "60% LTV / 82.5% LTC / 65% LTARV", "N/A"],
        ],
    )

    # Notes
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2,
        value="*90% LTC requires budget < $500K; otherwise 85%").font = FONT_NOTE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2,
        value="Loan Range: $100K-$3.5M | Terms: 6-24mo | Min FICO: 680 | Excluded: IL").font = FONT_NOTE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2,
        value="ZHVI >200%: -5% leverage adj. | ZHVI >300%: -10% leverage adj.").font = FONT_NOTE

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    return ws


# ============================================================================
# FIDELIS LEVERAGE (Human-Readable Reference)
# ============================================================================

def build_fidelis_leverage_sheet(wb):
    ws = wb.create_sheet("Fidelis Leverage")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 3

    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value="FIDELIS INVESTORS \u2014 LEVERAGE GUIDELINES")
    c.font = FONT_REF_HEADER
    c.fill = FILL_DEEP_BLUE
    c.alignment = ALIGN_CENTER
    for cc in range(2, 8):
        ws.cell(row=1, column=cc).fill = FILL_DEEP_BLUE

    def _grid(start_row, title, col_headers, data_rows, note=""):
        r = start_row
        span = len(col_headers)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + span - 1)
        c = ws.cell(row=r, column=2, value=title)
        c.font = FONT_REF_SECTION
        c.fill = FILL_LIGHT_GRAY
        c.alignment = ALIGN_LEFT
        for cc in range(3, 2 + span):
            ws.cell(row=r, column=cc).fill = FILL_LIGHT_GRAY
        r += 1

        if note:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2 + span - 1)
            ws.cell(row=r, column=2, value=note).font = FONT_NOTE
            r += 1

        for i, h in enumerate(col_headers):
            c = ws.cell(row=r, column=2 + i, value=h)
            c.font = FONT_REF_BOLD
            c.fill = FILL_POWDER
            c.border = THIN_BORDER
            c.alignment = ALIGN_CENTER
        r += 1

        for row_data in data_rows:
            for i, val in enumerate(row_data):
                c = ws.cell(row=r, column=2 + i, value=val)
                c.font = FONT_REF
                c.border = THIN_BORDER
                c.alignment = ALIGN_CENTER if i > 0 else ALIGN_LEFT
            r += 1
        return r + 1

    row = 3
    row = _grid(row,
        "NATIONAL \u2014 FIX & FLIP / BRIDGE (excl. FL, CA, NY)",
        ["FICO", "Exp 5+", "Exp 3-4", "Exp 1-2", "Exp 0"],
        [
            ["760+",    "90% / 90% / 75%", "87.5% / 87.5% / 72.5%", "85% / 85% / 70%", "82.5% / 82.5% / 67.5%"],
            ["740-759", "87.5% / 87.5% / 72.5%", "85% / 85% / 70%", "82.5% / 82.5% / 67.5%", "80% / 80% / 65%"],
            ["720-739", "85% / 85% / 70%", "82.5% / 82.5% / 67.5%", "80% / 80% / 65%", "77.5% / 77.5% / 62.5%"],
            ["700-719", "82.5% / 82.5% / 67.5%", "80% / 80% / 65%", "77.5% / 77.5% / 62.5%", "75% / 75% / 60%"],
            ["680-699", "80% / 80% / 65%", "77.5% / 77.5% / 62.5%", "75% / 75% / 60%", "72.5% / 72.5% / 57.5%"],
            ["660-679", "77.5% / 77.5% / 62.5%", "75% / 75% / 60%", "N/A", "N/A"],
        ],
        note="Format: Max LTV / Max LTC / Max LTARV"
    )

    row = _grid(row,
        "GUC \u2014 NATIONAL",
        ["FICO", "Exp 5+", "Exp 3-4", "Exp 1-2"],
        [
            ["740+",    "87.5% LTC / 72.5% LTARV", "85% LTC / 70% LTARV", "82.5% LTC / 67.5% LTARV"],
            ["720-739", "85% LTC / 70% LTARV", "82.5% LTC / 67.5% LTARV", "80% LTC / 65% LTARV"],
            ["700-719", "82.5% LTC / 67.5% LTARV", "80% LTC / 65% LTARV", "77.5% LTC / 62.5% LTARV"],
            ["680-699", "80% LTC / 65% LTARV", "77.5% LTC / 62.5% LTARV", "75% LTC / 60% LTARV"],
        ],
    )

    # Notes
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2,
        value="Loan Range: $75K-$5M | Terms: 6-24mo | Min FICO: 660 (Tier 1-2)").font = FONT_NOTE
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    ws.cell(row=row, column=2,
        value="All states eligible | Prepay: None | Extension: 0.5-1.0pt").font = FONT_NOTE

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    return ws


# ============================================================================
# ZILLOW MARKET DATA (copy from existing file)
# ============================================================================

def copy_zillow_data(wb, existing_path):
    print("  Reading existing Zillow Market Data ...")
    src = load_workbook(existing_path, read_only=True, data_only=True)
    src_ws = src["Zillow Market Data"]

    ws = wb.create_sheet("Zillow Market Data")

    total_rows = src_ws.max_row
    total_cols = src_ws.max_column

    print(f"  Copying {total_rows:,} rows x {total_cols} columns ...")

    for row_idx, row in enumerate(src_ws.iter_rows(min_row=1, max_row=total_rows,
                                                    max_col=total_cols,
                                                    values_only=True), start=1):
        for col_idx, val in enumerate(row, start=1):
            if val is None:
                continue
            new_cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if row_idx == 1:
                new_cell.font = Font(name="Calibri", size=9, bold=True, color=WHITE)
                new_cell.fill = FILL_DEEP_BLUE
                new_cell.alignment = ALIGN_CENTER
                new_cell.border = THIN_BORDER
            else:
                new_cell.font = Font(name="Calibri", size=9, color=DARK_TEXT)
                if col_idx >= 10:
                    new_cell.number_format = '$#,##0'

        if row_idx % 5000 == 0:
            print(f"    ... {row_idx:,} / {total_rows:,} rows")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{total_rows}"

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 36
    ws.column_dimensions["I"].width = 20

    src.close()
    print("  Zillow Market Data copied successfully.")
    return ws


# ============================================================================
# MAIN BUILD FUNCTION
# ============================================================================

def build_sizer_workbook():
    print("=" * 60)
    print("A&S Capital Sizer -- Excel Template Builder (v4)")
    print("Eastview Architecture + Colchis Leverage/Pricing")
    print("=" * 60)

    os.makedirs(ASSETS_DIR, exist_ok=True)

    # ------------------------------------------------------------------
    # Step 1: Check for existing Zillow data
    # ------------------------------------------------------------------
    has_zillow = False
    if os.path.exists(EXISTING_PATH):
        try:
            test_wb = load_workbook(EXISTING_PATH, read_only=True)
            if "Zillow Market Data" in test_wb.sheetnames:
                has_zillow = True
                print(f"[OK] Found existing Zillow Market Data in {EXISTING_PATH}")
            test_wb.close()
        except Exception as e:
            print(f"[WARN] Could not read existing file: {e}")

    # ------------------------------------------------------------------
    # Step 2: Create new workbook
    # ------------------------------------------------------------------
    wb = Workbook()

    print("\n[1/5] Building Sizer sheet (inputs + auto-sizing) ...")
    build_sizer_sheet(wb)

    print("[2/5] Building Colchis Data (hidden lookup) ...")
    build_colchis_data_sheet(wb)

    print("[3/5] Building Colchis Leverage (reference) ...")
    build_colchis_leverage_sheet(wb)

    print("[4/5] Building Fidelis Leverage (reference) ...")
    build_fidelis_leverage_sheet(wb)

    print("[5/5] Copying Zillow Market Data ...")
    if has_zillow:
        copy_zillow_data(wb, EXISTING_PATH)
    else:
        ws = wb.create_sheet("Zillow Market Data")
        ws.cell(row=1, column=1, value="RegionName")
        ws.cell(row=1, column=2, value="ZHVI")
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        print("  [WARN] No existing Zillow data found. Created placeholder sheet.")

    # ------------------------------------------------------------------
    # Step 3: Set active sheet to Sizer
    # ------------------------------------------------------------------
    wb.active = wb.sheetnames.index("Sizer")

    # ------------------------------------------------------------------
    # Step 4: Save
    # ------------------------------------------------------------------
    print(f"\nSaving workbook to: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    file_size = os.path.getsize(OUTPUT_PATH)
    print(f"[DONE] File saved successfully ({file_size / 1024 / 1024:.1f} MB)")
    print("=" * 60)


# ============================================================================
if __name__ == "__main__":
    build_sizer_workbook()
