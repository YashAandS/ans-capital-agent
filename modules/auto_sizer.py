"""
modules/auto_sizer.py
Automatic Sizer Filler — accepts multiple deal documents (PDFs, Excel files),
extracts all deal data using Claude AI, determines the loan type, fills the
correct sizer template, and highlights unfilled cells in red.

Supports any combination of:
  - Appraisals, loan applications, broker deal sheets (.xlsx, .xls, .pdf)
  - Credit reports, term sheets, HUDs, intake forms
  - Any document with deal information
"""

import io
import json
import os
import re
from datetime import datetime

import openpyxl
import pdfplumber
from anthropic import Anthropic

from modules.sizer import get_template_path, SIZER_MAPS, SIZER_TEMPLATES, _extract_x14_blocks, _patch_x14_into_output


def _lookup_zip_code(api_key: str, address: str, city: str, state: str) -> str:
    """Use Claude (Sonnet) to find a ZIP code via web-style reasoning."""
    if not address or not city:
        return ""
    client = Anthropic(api_key=api_key)
    query = f"{address}, {city}, {state}" if state else f"{address}, {city}"
    response = client.messages.create(
        model="claude-4-sonnet-20250514",
        max_tokens=20,
        system="You are a US ZIP code lookup tool. Given a property address, respond with ONLY the 5-digit ZIP code. Nothing else. If you cannot determine it, respond with UNKNOWN.",
        messages=[{"role": "user", "content": f"What is the ZIP code for: {query}"}],
    )
    result = response.content[0].text.strip()
    match = re.search(r"\b(\d{5})\b", result)
    return int(match.group(1)) if match else ""


# ---------------------------------------------------------------------------
# Cells with data-validation (dropdowns) that must NOT be overwritten.
# These are critical for the sizer template algorithms / scoring logic.
# Format: { (sheet_name, cell_ref), ... }
# ---------------------------------------------------------------------------
DROPDOWN_CELLS = {
    # --- RTL ---
    # Property type (H57-H61), Change of use (W57-W61)
    ("Sizer", "H57"), ("Sizer", "H58"), ("Sizer", "H59"), ("Sizer", "H60"), ("Sizer", "H61"),
    ("Sizer", "W57"), ("Sizer", "W58"), ("Sizer", "W59"), ("Sizer", "W60"), ("Sizer", "W61"),
    # Is guarantor? Yes/No (F26, F31, F36, F41)
    ("Sizer", "F26"), ("Sizer", "F31"), ("Sizer", "F36"), ("Sizer", "F41"),
    # Closing Proceeds dropdown (E15)
    ("Sizer", "E15"),
    # Loan Program (G15), Loan Term (G16), Interest Accrual (G17), Prepayment Penalty (G26)
    ("Summary", "G15"), ("Summary", "G16"), ("Summary", "G17"), ("Summary", "G26"),

    # --- DSCR ---
    # Amortization (R126), Rate Type (R127), Low Perf Market (R128), Rate (R129)
    ("Sizer", "R126"), ("Sizer", "R127"), ("Sizer", "R128"), ("Sizer", "R129"),
    # Prepayment Penalty (R124)
    ("Sizer", "R124"),
    # Origination fee tier (G117), Property type (G118)
    ("Sizer", "G117"), ("Sizer", "G118"),
    # Closing Proceeds dropdown (G20)
    ("Sizer", "G20"),
    # NOTE: DSCR FICO cells G35/G41/G47/G53 are dropdowns but we allow writing
    # because "Foreign National" is a valid dropdown value in the Pricing sheet.
    # Condo warrantability Yes/No (L71-L78)
    ("Sizer", "L71"), ("Sizer", "L72"), ("Sizer", "L73"), ("Sizer", "L74"),
    ("Sizer", "L75"), ("Sizer", "L76"), ("Sizer", "L77"), ("Sizer", "L78"),
    # Property sheet dropdowns
    ("Property", "H5"), ("Property", "I5"), ("Property", "B5"),
    ("Property", "AP5"), ("Property", "AV5"), ("Property", "BB5"),

    # --- MF ---
    # Closing Proceeds (G24), Loan Program (G25), Loan Term (G26), Exit Strategy (G27)
    ("Sizer", "G24"), ("Sizer", "G25"), ("Sizer", "G26"), ("Sizer", "G27"),
    # NOTE: State (G16) is a dropdown but we allow writing — AI provides valid 2-letter codes
    # Change of use (T23)
    ("Sizer", "T23"),
    # Leverage deductions Yes/No (G84, G85)
    ("Sizer", "G84"), ("Sizer", "G85"),

    # --- GUC ---
    # Closing Proceeds (G25), Loan Term (G26)
    ("Sizer", "G25"), ("Sizer", "G26"),
    # State abbreviation dropdown (G16) — same cell as MF
    # Leverage deductions Yes/No (G58, G59, G60)
    ("Sizer", "G58"), ("Sizer", "G59"), ("Sizer", "G60"),
    # Num units dropdown (G19)
    ("Sizer", "G19"),
}

# Fields to skip when highlighting (these are optional / not always applicable)
OPTIONAL_FIELDS = {
    # RTL optional
    "closing_proceeds", "loan_id", "exception_pricing", "exception_reasoning",
    "interest_accrual_type",
    "guarantor_2_name", "guarantor_2_fico", "guarantor_2_credit_date",
    "guarantor_2_is_guarantor", "guarantor_2_ownership",
    "guarantor_3_name", "guarantor_3_fico", "guarantor_3_credit_date",
    "guarantor_3_is_guarantor", "guarantor_3_ownership",
    "guarantor_4_name", "guarantor_4_fico", "guarantor_4_credit_date",
    "guarantor_4_is_guarantor", "guarantor_4_ownership",
    "prop1_secondary_aiv", "prop1_secondary_arv", "prop1_completed_rehab",
    "prop1_change_of_use", "g1_gc_not_owner",
    # DSCR optional
    "condo_not_completed", "condo_hoa_not_turned", "condo_single_owner_10",
    "condo_short_term", "condo_majority_rental", "condo_litigation",
    "condo_commercial_25", "condo_arrears_15",
    "guarantor_2_first", "guarantor_2_last", "guarantor_2_fico",
    "guarantor_2_credit_date", "guarantor_2_ownership",
    "guarantor_3_first", "guarantor_3_last", "guarantor_3_fico",
    "guarantor_3_credit_date", "guarantor_3_ownership",
    "guarantor_4_first", "guarantor_4_last", "guarantor_4_fico",
    "guarantor_4_credit_date", "guarantor_4_ownership",
    "cash_out_amount", "broker_orig_pct", "buydown_pct", "low_perf_market",
    "escrow_taxes", "escrow_hazard", "escrow_flood",
    "prop_annual_flood_ins", "prop_annual_hoa",  # flood & HOA not always applicable
    # MF optional
    "guarantor_2_name", "guarantor_2_fico", "guarantor_2_credit_date",
    "guarantor_2_ownership", "completed_rehab", "change_of_use",
    "declining_rents", "declining_hp_appraisal",
    "exit_dscr_rate", "exit_agency_rate", "exit_agency_advance", "exit_agency_dscr",
    "lender_orig_pct", "broker_orig_pct",
    # GUC optional
    "completed_rehab",
    # Rate fields (user fills in Excel)
    "loan_interest_rate", "borrower_pricing",
    # Dropdown fields the user sets themselves in Excel
    "closing_proceeds", "loan_term", "loan_program",
    "interest_accrual_type", "prepay_option",
}


# ---------------------------------------------------------------------------
# Field definitions for each loan type (used in the AI prompt)
# ---------------------------------------------------------------------------

def _get_field_descriptions(loan_type: str) -> dict:
    """Return human-readable descriptions for each sizer field."""

    COMMON_FIELDS = {
        "closing_date": "Expected closing date (YYYY-MM-DD)",
        "entity_name": "Borrower entity name (LLC, Corp, etc.)",
        "loan_id": "Loan ID or reference number",
        "initial_loan_amount": "Initial loan amount in dollars",
        "interest_reserves": "Financed interest reserves in dollars",
        "verified_liquidity": "Verified borrower liquidity in dollars",
    }

    RTL_FIELDS = {
        **COMMON_FIELDS,
        "num_owners": "Number of owners/guarantors (1-4)",
        "guarantor_1_name": "Guarantor 1 full name",
        "guarantor_1_fico": "Guarantor 1 FICO credit score (integer)",
        "guarantor_1_credit_date": "Guarantor 1 credit report date (YYYY-MM-DD)",
        "guarantor_1_is_guarantor": "Is guarantor? (Yes or No)",
        "guarantor_1_ownership": "Guarantor 1 ownership percentage as decimal (e.g. 1.0 = 100%)",
        "guarantor_2_name": "Guarantor 2 full name (if applicable)",
        "guarantor_2_fico": "Guarantor 2 FICO score (if applicable)",
        "guarantor_2_credit_date": "Guarantor 2 credit report date (if applicable)",
        "guarantor_2_is_guarantor": "Guarantor 2 is guarantor? (Yes or No)",
        "guarantor_2_ownership": "Guarantor 2 ownership percentage as decimal",
        "g1_rehab_sold": "# rehab projects completed & sold",
        "g1_rehab_refinanced": "# rehab projects completed & refinanced as rental",
        "g1_acquired_rental": "# properties acquired as rental",
        "g1_gc_not_owner": "# general contractor projects (not owner)",
        "prop1_address": "Property street address",
        "prop1_city": "Property city",
        "prop1_state": "Property state (2-letter)",
        "prop1_zip": "Property ZIP code (integer)",
        "prop1_type": "Property type (SFR, Townhome, Condo, PUD, 2 Unit, 3 Unit, 4 Unit)",
        "prop1_appraisal_date": "Appraisal date (YYYY-MM-DD)",
        "prop1_as_is_value": "As-is appraised value in dollars",
        "prop1_secondary_aiv": "Secondary as-is value in dollars (if available)",
        "prop1_arv": "After-repair value in dollars",
        "prop1_secondary_arv": "Secondary ARV in dollars (if available)",
        "prop1_rehab_budget": "Rehab budget in dollars",
        "prop1_pre_rehab_sqft": "Pre-rehab square footage (integer)",
        "prop1_post_rehab_sqft": "Post-rehab square footage (integer)",
        "prop1_purchase_date": "Purchase date (YYYY-MM-DD)",
        "prop1_purchase_price": "Purchase price in dollars",
        "prop1_change_of_use": "Change of use? (Yes or No)",
        "loan_program": "Loan program (Fix & Flip, Bridge, Bridge Plus)",
        "loan_term": "Loan term (e.g. '12 Months', '18 Months', '24 Months', '36 Months')",
        "financed_rehab": "Financed rehab budget in dollars",
    }

    DSCR_FIELDS = {
        **COMMON_FIELDS,
        "interest_reserves": None,  # Not in DSCR
        "prop_purchase_date": "Property purchase date (YYYY-MM-DD)",
        "num_guarantors": "Number of guarantors (1-4)",
        "guarantor_1_first": "Guarantor 1 first name",
        "guarantor_1_last": "Guarantor 1 last name",
        "guarantor_1_fico": "Guarantor 1 FICO score (integer)",
        "guarantor_1_credit_date": "Guarantor 1 credit report date (YYYY-MM-DD)",
        "guarantor_1_ownership": "Guarantor 1 ownership percentage as decimal",
        "guarantor_2_first": "Guarantor 2 first name (if applicable)",
        "guarantor_2_last": "Guarantor 2 last name (if applicable)",
        "guarantor_2_fico": "Guarantor 2 FICO score (if applicable)",
        "guarantor_2_credit_date": "Guarantor 2 credit report date (if applicable)",
        "guarantor_2_ownership": "Guarantor 2 ownership percentage as decimal (if applicable)",
        "property_type": "Predominant property type for loan structure (SFR, Townhome, Condo, etc.)",
        "amortization": "Amortization type (Fully Amortizing or Interest Only)",
        "rate_type": "Rate type (FIXED 30, 5/1 ARM, 7/1 ARM)",
        "verified_liquidity": "Verified borrower liquidity in dollars",
        # Property details (Property sheet)
        "prop_address": "Property street address",
        "prop_city": "Property city",
        "prop_state": "Property state (2-letter)",
        "prop_zip": "Property ZIP code (integer)",
        "prop_type": "Property type (SFR, Townhome, Condo, PUD, 2 Unit, 3 Unit, 4 Unit)",
        "prop_sqft": "Property square footage (integer)",
        "prop_num_units": "Number of units (integer)",
        "prop_appraisal_date": "Appraisal date (YYYY-MM-DD)",
        "prop_appraisal_value": "Appraisal as-is value in dollars",
        "prop_purchase_price": "Purchase price in dollars",
        # Rent & Expenses
        "prop_monthly_rent": "Monthly rent in place in dollars",
        "prop_market_rent": "Monthly market rent in dollars",
        "prop_annual_taxes": "Annual property taxes in dollars",
        "prop_annual_hazard_ins": "Annual hazard insurance in dollars",
        "prop_annual_flood_ins": "Annual flood insurance in dollars",
        "prop_annual_hoa": "Annual HOA fees in dollars",
    }
    DSCR_FIELDS = {k: v for k, v in DSCR_FIELDS.items() if v is not None}

    MF_FIELDS = {
        **COMMON_FIELDS,
        "address": "Property address",
        "city": "City",
        "state": "State (2-letter)",
        "zip_code": "ZIP code (integer)",
        "num_units": "Number of units (integer, must be 5+)",
        "loan_program": "Loan program (Bridge, CAPEX)",
        "loan_term": "Loan term (e.g. '12 Months')",
        "purchase_price": "Purchase price in dollars",
        "purchase_date": "Purchase date (YYYY-MM-DD)",
        "appraisal_date": "Appraisal date (YYYY-MM-DD)",
        "as_is_value": "As-is value in dollars",
        "arv": "After-repair value / ARV in dollars",
        "rehab_budget": "Rehab budget in dollars",
        "pre_rehab_sqft": "Pre-rehab square footage",
        "post_rehab_sqft": "Post-rehab square footage",
        "gross_potential_rev": "Annual gross potential revenue in dollars",
        "opex_vacancy": "Annual operating expenses & vacancy in dollars",
        "annual_taxes": "Annual taxes in dollars",
        "annual_insurance": "Annual insurance in dollars",
        "num_owners": "Number of owners",
        "guarantor_1_name": "Guarantor 1 full name",
        "guarantor_1_fico": "Guarantor 1 FICO score",
        "guarantor_1_credit_date": "Guarantor 1 credit report date (YYYY-MM-DD)",
        "guarantor_1_ownership": "Guarantor 1 ownership percentage as decimal",
        "financed_rehab": "Financed rehab in dollars",
    }

    GUC_FIELDS = {
        **COMMON_FIELDS,
        "address": "Property address",
        "city": "City",
        "state": "State (2-letter)",
        "zip_code": "ZIP code (integer)",
        "num_units": "Number of units",
        "loan_term": "Loan term (e.g. '12 Months')",
        "purchase_price": "Purchase price in dollars",
        "purchase_date": "Purchase date (YYYY-MM-DD)",
        "appraisal_date": "Appraisal date (YYYY-MM-DD)",
        "as_is_value": "As-is value in dollars",
        "arv": "After-repair value / ARV in dollars",
        "rehab_budget": "Construction budget in dollars",
        "post_completion_sqft": "Post-completion square footage",
        "num_owners": "Number of owners",
        "guarantor_1_name": "Guarantor 1 full name",
        "guarantor_1_fico": "Guarantor 1 FICO score",
        "guarantor_1_credit_date": "Guarantor 1 credit report date (YYYY-MM-DD)",
        "guarantor_1_ownership": "Guarantor 1 ownership percentage as decimal",
        "g1_construction_completed": "Total # GUC projects completed (sold + rented combined)",
        "entitled_land": "Entitled land? (Yes or No)",
        "approved_permits": "Approved permits & plans? (Yes or No)",
        "interest_reserves_flag": "Interest reserves? (Yes or No)",
    }

    return {
        "RTL": RTL_FIELDS,
        "DSCR": DSCR_FIELDS,
        "MF": MF_FIELDS,
        "GUC": GUC_FIELDS,
    }.get(loan_type, {})


# ---------------------------------------------------------------------------
# Extract text from uploaded files
# ---------------------------------------------------------------------------

def extract_text_from_excel(file_bytes: bytes) -> str:
    """Read all sheets from an Excel file and return as structured text."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            row_vals = []
            for cell in row:
                if cell.value is not None:
                    row_vals.append(f"{cell.coordinate}: {cell.value}")
            if row_vals:
                lines.append("  ".join(row_vals))
    return "\n".join(lines)


def extract_text_from_pdf(file_bytes: bytes) -> str:
    """Read all pages from a PDF and return as text."""
    pages = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text:
                pages.append(f"--- Page {i} ---\n{text}")
    return "\n".join(pages)


def extract_text_from_file(file_bytes: bytes, file_name: str) -> str:
    """Extract text from a file based on its extension."""
    ext = os.path.splitext(file_name)[1].lower()
    if ext in (".xlsx", ".xls"):
        return extract_text_from_excel(file_bytes)
    elif ext == ".pdf":
        return extract_text_from_pdf(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Upload .xlsx or .pdf files.")


def extract_text_from_multiple_files(files: list) -> str:
    """
    Extract text from multiple uploaded files and combine them.
    Each item in files should be a dict with 'bytes' and 'name' keys.
    """
    all_text = []
    for i, f in enumerate(files, 1):
        try:
            text = extract_text_from_file(f["bytes"], f["name"])
            all_text.append(f"\n{'='*60}\nDOCUMENT {i}: {f['name']}\n{'='*60}\n{text}")
        except Exception as e:
            all_text.append(f"\n[Could not read {f['name']}: {e}]")
    return "\n".join(all_text)


# ---------------------------------------------------------------------------
# AI-powered extraction
# ---------------------------------------------------------------------------

def detect_loan_type(api_key: str, document_text: str) -> str:
    """Use Claude to detect the loan type from the combined documents."""
    client = Anthropic(api_key=api_key)

    response = client.messages.create(
        model="claude-4-sonnet-20250514",
        max_tokens=50,
        system="You are a loan type classifier for a private lending company. Based on the document content, determine the loan type. Respond with ONLY one of these four values: RTL, DSCR, MF, GUC. Nothing else.",
        messages=[{
            "role": "user",
            "content": f"""Classify this loan into one of these types based on the documents:
- RTL: Fix & Flip / Bridge loans (short-term rehab loans for 1-4 unit residential)
- DSCR: Debt Service Coverage Ratio rental loans (long-term rental, 30-year, 1-4 units)
- MF: Multifamily (5+ unit apartment buildings)
- GUC: Ground Up Construction (building from scratch on vacant land)

Document content:
{document_text[:10000]}

Respond with ONLY the loan type code (RTL, DSCR, MF, or GUC):"""
        }]
    )

    result = response.content[0].text.strip().upper()
    if result in ("RTL", "DSCR", "MF", "GUC"):
        return result
    for lt in ("RTL", "DSCR", "MF", "GUC"):
        if lt in result:
            return lt
    return "RTL"


def extract_fields_from_documents(api_key: str, document_text: str, loan_type: str) -> dict:
    """
    Use Claude to extract sizer fields from multiple combined documents.
    Returns a dict of { field_name: value }.
    """
    client = Anthropic(api_key=api_key)

    field_descriptions = _get_field_descriptions(loan_type)
    fields_prompt = "\n".join(
        f'  "{k}": {v}' for k, v in field_descriptions.items()
    )

    # Truncate to fit context — leave room for the prompt
    max_doc_chars = 150000
    if len(document_text) > max_doc_chars:
        document_text = document_text[:max_doc_chars] + "\n... [truncated]"

    response = client.messages.create(
        model="claude-4-opus-20250514",
        max_tokens=4096,
        system="""You are a data extraction specialist for A&S Capital, a private lending company.
You are given MULTIPLE documents for a single loan deal (could be appraisals, loan applications, credit reports, broker sheets, term sheets, etc.).

Your job: cross-reference ALL the documents and extract every piece of deal information you can find into a structured JSON object.

RULES:
- Return ONLY valid JSON — no markdown, no explanation, no code fences
- Use the exact field names provided
- For dollar amounts: return as numbers (no $ sign, no commas). Example: 500000
- For percentages/rates: return as decimals. Example: 0.75 for 75%, 0.02 for 2%
- For dates: return as "YYYY-MM-DD" strings
- For integers (FICO, units, ZIP): return as numbers
- For ownership: return as decimal (1.0 = 100%, 0.5 = 50%)
- If a field is not found in ANY of the documents, omit it entirely
- For loan term, return as text like "12 Months", "18 Months", etc.
- For Yes/No fields, return "Yes" or "No"
- If the guarantor name is a single full name and the loan type is DSCR (which needs first/last separately), split it
- If documents have conflicting values, prefer the most recent or most authoritative source (appraisal > broker sheet)
- Extract EVERYTHING you can find — the more fields you fill, the better
- IMPORTANT: If the borrower/guarantor is a Foreign National (non-US citizen without a US credit history), set their FICO field to the string "Foreign National" instead of a number""",
        messages=[{
            "role": "user",
            "content": f"""Extract the following fields for a {loan_type} loan from these documents.

FIELDS TO EXTRACT:
{fields_prompt}

DOCUMENTS:
{document_text}

Return ONLY a JSON object with the extracted fields:"""
        }]
    )

    raw = response.content[0].text.strip()

    # Clean up markdown fences
    if raw.startswith("```"):
        raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
    if raw.endswith("```"):
        raw = raw[:-3]
    if raw.startswith("json"):
        raw = raw[4:]
    raw = raw.strip()

    try:
        extracted = json.loads(raw)
    except json.JSONDecodeError:
        start = raw.find("{")
        end = raw.rfind("}") + 1
        if start >= 0 and end > start:
            extracted = json.loads(raw[start:end])
        else:
            extracted = {}

    return extracted


# ---------------------------------------------------------------------------
# Fill sizer with red highlighting for missing fields
# ---------------------------------------------------------------------------

def fill_sizer_with_highlights(
    template_path: str,
    loan_type: str,
    extracted: dict,
) -> tuple:
    """
    Fill the sizer template with extracted data.
    Skips cells with data-validation dropdowns to preserve template integrity.
    Returns a list of missing fields (no red highlighting).

    Returns:
        (output: BytesIO, filled_count: int, missing_fields: list[str])
    """
    input_map = SIZER_MAPS.get(loan_type)
    if not input_map:
        raise ValueError(f"Unknown loan type: {loan_type}")

    # Step 1: Extract x14 dropdown blocks BEFORE openpyxl touches the file
    x14_blocks, ws_tags = _extract_x14_blocks(template_path)

    # Step 2: Load workbook and fill cells
    wb = openpyxl.load_workbook(template_path)

    filled_count = 0
    missing_fields = []

    for field_name, (sheet_name, cell_ref) in input_map.items():
        if sheet_name not in wb.sheetnames:
            continue

        # Skip cells with dropdown data validation — never overwrite these
        if (sheet_name, cell_ref) in DROPDOWN_CELLS:
            continue

        ws = wb[sheet_name]
        value = extracted.get(field_name)

        if value is not None and value != "":
            ws[cell_ref] = value
            filled_count += 1
        else:
            # Track missing fields (skip optional ones)
            if field_name in OPTIONAL_FIELDS:
                continue
            missing_fields.append(field_name)

    # Step 3: Save (openpyxl will strip x14 dropdowns)
    output = io.BytesIO()
    wb.save(output)

    # Step 4: Patch x14 dropdowns back in at ZIP level
    output = _patch_x14_into_output(output, x14_blocks, ws_tags)
    return output, filled_count, missing_fields


# ---------------------------------------------------------------------------
# Main auto-sizer function
# ---------------------------------------------------------------------------

def auto_fill_sizer(
    api_key: str,
    assets_dir: str,
    files: list,
    loan_type_override: str = None,
) -> tuple:
    """
    Full pipeline: read multiple docs → detect loan type → extract fields → fill sizer with highlights.

    Args:
        api_key: Anthropic API key
        assets_dir: Path to assets directory
        files: List of dicts with 'bytes' and 'name' keys
        loan_type_override: If provided, skip loan type detection

    Returns:
        (sizer_bytes: BytesIO, loan_type: str, filled_count: int, missing_fields: list, extracted_fields: dict)
    """

    # Step 1: Extract text from all uploaded files
    document_text = extract_text_from_multiple_files(files)

    # Step 2: Detect or use provided loan type
    if loan_type_override and loan_type_override in SIZER_MAPS:
        loan_type = loan_type_override
    else:
        loan_type = detect_loan_type(api_key, document_text)

    # Step 3: Extract fields using AI across all documents
    extracted = extract_fields_from_documents(api_key, document_text, loan_type)

    # Step 4: Post-processing — auto-fill dates, ZIP lookup, Foreign National
    today_str = datetime.now().strftime("%Y-%m-%d")

    # --- Auto-fill appraisal date and credit report date with today ---
    APPRAISAL_DATE_FIELDS = {
        "prop1_appraisal_date", "prop_appraisal_date", "appraisal_date",
    }
    CREDIT_DATE_FIELDS = {
        "guarantor_1_credit_date", "guarantor_2_credit_date",
        "guarantor_3_credit_date", "guarantor_4_credit_date",
    }
    input_map = SIZER_MAPS.get(loan_type, {})
    for field in APPRAISAL_DATE_FIELDS:
        if field in input_map and not extracted.get(field):
            extracted[field] = today_str
    for field in CREDIT_DATE_FIELDS:
        if field in input_map and not extracted.get(field):
            # Only fill if the corresponding guarantor exists
            gnum = field.split("_")[1]  # "1", "2", etc.
            name_keys = [f"guarantor_{gnum}_name", f"guarantor_{gnum}_first"]
            if any(extracted.get(k) for k in name_keys):
                extracted[field] = today_str

    # --- Foreign National: write "Foreign National" to FICO cells ---
    FICO_FIELDS = {
        "guarantor_1_fico", "guarantor_2_fico",
        "guarantor_3_fico", "guarantor_4_fico",
    }
    for field in FICO_FIELDS:
        val = extracted.get(field)
        if isinstance(val, str) and "foreign" in val.lower():
            extracted[field] = "Foreign National"

    # --- ZIP code lookup if missing ---
    ZIP_FIELD_SETS = {
        "RTL": ("prop1_zip", "prop1_address", "prop1_city", "prop1_state"),
        "DSCR": ("prop_zip", "prop_address", "prop_city", "prop_state"),
        "MF": ("zip_code", "address", "city", "state"),
        "GUC": ("zip_code", "address", "city", "state"),
    }
    if loan_type in ZIP_FIELD_SETS:
        zip_field, addr_field, city_field, state_field = ZIP_FIELD_SETS[loan_type]
        if not extracted.get(zip_field):
            addr = extracted.get(addr_field, "")
            city = extracted.get(city_field, "")
            state = extracted.get(state_field, "")
            if addr and city:
                try:
                    zip_val = _lookup_zip_code(api_key, addr, city, state)
                    if zip_val:
                        extracted[zip_field] = zip_val
                except Exception:
                    pass  # Non-critical — skip if lookup fails

    # Step 5: Fill the sizer template
    template_path = get_template_path(assets_dir, loan_type)
    sizer_bytes, filled_count, missing_fields = fill_sizer_with_highlights(
        template_path, loan_type, extracted
    )

    return sizer_bytes, loan_type, filled_count, missing_fields, extracted
