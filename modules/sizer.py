"""
modules/sizer.py
Excel Sizer Filler — opens the actual Eastview sizer templates and fills in
user-input cells. The workbooks contain formulas, VLOOKUPs to Zillow data,
leverage grids, and scoring logic that compute everything automatically.

Supported sizer types: RTL, DSCR, MF (Multifamily 5+), GUC (Ground Up Construction)
"""

import io
import os
from datetime import datetime
from copy import copy
import openpyxl


# ---------------------------------------------------------------------------
# Cell mapping per sizer type
# Each map defines: { field_name: (sheet_name, cell_coordinate) }
# Only USER-INPUT cells are mapped — formulas are left intact.
# ---------------------------------------------------------------------------

RTL_INPUT_MAP = {
    # === Sizer Sheet: Loan Purpose ===
    "closing_date":           ("Sizer", "E12"),
    "closing_proceeds":       ("Sizer", "E15"),
    # === Sizer Sheet: Borrower / Entity ===
    "entity_name":            ("Sizer", "F21"),
    "num_owners":             ("Sizer", "F22"),
    "guarantor_1_name":       ("Sizer", "F23"),
    "guarantor_1_fico":       ("Sizer", "F24"),
    "guarantor_1_credit_date":("Sizer", "F25"),
    "guarantor_1_is_guarantor":("Sizer", "F26"),
    "guarantor_1_ownership":  ("Sizer", "F27"),
    "guarantor_2_name":       ("Sizer", "F28"),
    "guarantor_2_fico":       ("Sizer", "F29"),
    "guarantor_2_credit_date":("Sizer", "F30"),
    "guarantor_2_is_guarantor":("Sizer", "F31"),
    "guarantor_2_ownership":  ("Sizer", "F32"),
    "guarantor_3_name":       ("Sizer", "F33"),
    "guarantor_3_fico":       ("Sizer", "F34"),
    "guarantor_3_credit_date":("Sizer", "F35"),
    "guarantor_3_is_guarantor":("Sizer", "F36"),
    "guarantor_3_ownership":  ("Sizer", "F37"),
    "guarantor_4_name":       ("Sizer", "F38"),
    "guarantor_4_fico":       ("Sizer", "F39"),
    "guarantor_4_credit_date":("Sizer", "F40"),
    "guarantor_4_is_guarantor":("Sizer", "F41"),
    "guarantor_4_ownership":  ("Sizer", "F42"),
    # === Experience (Guarantor 1) ===
    "g1_rehab_sold":          ("Sizer", "F45"),
    "g1_rehab_refinanced":    ("Sizer", "F46"),
    "g1_acquired_rental":     ("Sizer", "F47"),
    "g1_gc_not_owner":        ("Sizer", "F48"),
    # === Property row 57 (property 1) ===
    "prop1_address":          ("Sizer", "D57"),
    "prop1_city":             ("Sizer", "E57"),
    "prop1_state":            ("Sizer", "F57"),
    "prop1_zip":              ("Sizer", "G57"),
    "prop1_type":             ("Sizer", "H57"),
    "prop1_appraisal_date":   ("Sizer", "N57"),
    "prop1_as_is_value":      ("Sizer", "O57"),
    "prop1_secondary_aiv":    ("Sizer", "P57"),
    "prop1_arv":              ("Sizer", "Q57"),
    "prop1_secondary_arv":    ("Sizer", "R57"),
    "prop1_completed_rehab":  ("Sizer", "S57"),
    "prop1_rehab_budget":     ("Sizer", "T57"),
    "prop1_pre_rehab_sqft":   ("Sizer", "U57"),
    "prop1_post_rehab_sqft":  ("Sizer", "V57"),
    "prop1_change_of_use":    ("Sizer", "W57"),
    "prop1_purchase_date":    ("Sizer", "X57"),
    "prop1_purchase_price":   ("Sizer", "Y57"),
    # === Summary Sheet overrides ===
    "loan_program":           ("Summary", "G15"),
    "loan_term":              ("Summary", "G16"),
    "interest_accrual_type":  ("Summary", "G17"),
    "initial_loan_amount":    ("Summary", "G62"),
    "interest_reserves":      ("Summary", "G64"),
    "financed_rehab":         ("Summary", "G66"),
    "loan_interest_rate":     ("Summary", "G80"),
    "loan_id":                ("Summary", "G84"),
    "exception_pricing":      ("Summary", "G87"),
    "exception_reasoning":    ("Summary", "K87"),
}

DSCR_INPUT_MAP = {
    # === Sizer Sheet: Loan Purpose ===
    "closing_date":           ("Sizer", "G16"),
    "prop_purchase_date":     ("Sizer", "G19"),
    # === Borrower / Entity ===
    "entity_name":            ("Sizer", "G30"),
    "num_guarantors":         ("Sizer", "G31"),
    "guarantor_1_first":      ("Sizer", "G32"),
    "guarantor_1_last":       ("Sizer", "G33"),
    "guarantor_1_fico":       ("Sizer", "G35"),
    "guarantor_1_credit_date":("Sizer", "G36"),
    "guarantor_1_ownership":  ("Sizer", "G37"),
    "guarantor_2_first":      ("Sizer", "G38"),
    "guarantor_2_last":       ("Sizer", "G39"),
    "guarantor_2_fico":       ("Sizer", "G41"),
    "guarantor_2_credit_date":("Sizer", "G42"),
    "guarantor_2_ownership":  ("Sizer", "G43"),
    "guarantor_3_first":      ("Sizer", "G44"),
    "guarantor_3_last":       ("Sizer", "G45"),
    "guarantor_3_fico":       ("Sizer", "G47"),
    "guarantor_3_credit_date":("Sizer", "G48"),
    "guarantor_3_ownership":  ("Sizer", "G49"),
    "guarantor_4_first":      ("Sizer", "G50"),
    "guarantor_4_last":       ("Sizer", "G51"),
    "guarantor_4_fico":       ("Sizer", "G53"),
    "guarantor_4_credit_date":("Sizer", "G54"),
    "guarantor_4_ownership":  ("Sizer", "G55"),
    # === Condo warrantability ===
    "condo_not_completed":    ("Sizer", "L71"),
    "condo_hoa_not_turned":   ("Sizer", "L72"),
    "condo_single_owner_10":  ("Sizer", "L73"),
    "condo_short_term":       ("Sizer", "L74"),
    "condo_majority_rental":  ("Sizer", "L75"),
    "condo_litigation":       ("Sizer", "L76"),
    "condo_commercial_25":    ("Sizer", "L77"),
    "condo_arrears_15":       ("Sizer", "L78"),
    # === Loan Structure (user fills rate, prepay, closing costs, requested amount themselves) ===
    "amortization":           ("Sizer", "R126"),
    "rate_type":              ("Sizer", "R127"),
    "low_perf_market":        ("Sizer", "R128"),
    "cash_out_amount":        ("Sizer", "V109"),
    "lender_orig_pct":        ("Sizer", "V110"),  # hardcoded to 2% in app.py
    "broker_orig_pct":        ("Sizer", "V112"),
    "buydown_pct":            ("Sizer", "V114"),
    "property_type":          ("Sizer", "G118"),
    "loan_id":                ("Sizer", "R137"),
    # === Property Sheet (first property, row 5) ===
    "prop_address":           ("Property", "D5"),
    "prop_city":              ("Property", "E5"),
    "prop_state":             ("Property", "F5"),
    "prop_zip":               ("Property", "G5"),
    "prop_type":              ("Property", "H5"),
    "prop_sqft":              ("Property", "J5"),
    "prop_num_units":         ("Property", "K5"),
    "prop_appraisal_date":    ("Property", "P5"),
    "prop_appraisal_value":   ("Property", "Q5"),
    "prop_purchase_price":    ("Property", "U5"),
    # === Property Rent (Unit 1, row 5) ===
    "prop_monthly_rent":      ("Property", "AQ5"),
    "prop_market_rent":       ("Property", "AR5"),
    # === Property Expenses (row 5) ===
    "prop_annual_taxes":      ("Property", "AE5"),
    "prop_annual_hazard_ins": ("Property", "AF5"),
    "prop_annual_flood_ins":  ("Property", "AG5"),
    "prop_annual_hoa":        ("Property", "AI5"),
    # === Liquidity ===
    "verified_liquidity":     ("Sizer", "V94"),
    # === Escrows ===
    "escrow_taxes":           ("Sizer", "V62"),
    "escrow_hazard":          ("Sizer", "V63"),
    "escrow_flood":           ("Sizer", "V64"),
}

MF_INPUT_MAP = {
    # === Property Information ===
    "address":                ("Sizer", "G14"),
    "city":                   ("Sizer", "G15"),
    "state":                  ("Sizer", "G16"),
    "zip_code":               ("Sizer", "G17"),
    "num_units":              ("Sizer", "G19"),
    # === Loan Information ===
    "closing_date":           ("Sizer", "G23"),
    "closing_proceeds":       ("Sizer", "G24"),
    "loan_program":           ("Sizer", "G25"),
    "loan_term":              ("Sizer", "G26"),
    # === Valuation & Rehab ===
    "purchase_price":         ("Sizer", "T14"),
    "purchase_date":          ("Sizer", "T15"),
    "appraisal_date":         ("Sizer", "T16"),
    "as_is_value":            ("Sizer", "T17"),
    "arv":                    ("Sizer", "T18"),
    "completed_rehab":        ("Sizer", "T19"),
    "rehab_budget":           ("Sizer", "T20"),
    "pre_rehab_sqft":         ("Sizer", "T21"),
    "post_rehab_sqft":        ("Sizer", "T22"),
    "change_of_use":          ("Sizer", "T23"),
    # === Property Economics ===
    "gross_potential_rev":    ("Sizer", "T26"),
    "opex_vacancy":           ("Sizer", "T27"),
    "annual_taxes":           ("Sizer", "T28"),
    "annual_insurance":       ("Sizer", "T29"),
    # === Borrower / Entity ===
    "entity_name":            ("Sizer", "G43"),
    "num_owners":             ("Sizer", "G44"),
    "guarantor_1_name":       ("Sizer", "G45"),
    "guarantor_1_fico":       ("Sizer", "G46"),
    "guarantor_1_credit_date":("Sizer", "G47"),
    "guarantor_1_ownership":  ("Sizer", "G48"),
    "guarantor_2_name":       ("Sizer", "G49"),
    "guarantor_2_fico":       ("Sizer", "G50"),
    "guarantor_2_credit_date":("Sizer", "G51"),
    "guarantor_2_ownership":  ("Sizer", "G52"),
    # === Experience (<=48 mo) ===
    "g1_rehab_sold":          ("Sizer", "G63"),
    "g1_rehab_retained":      ("Sizer", "G64"),
    "g1_acquired_rental":     ("Sizer", "G65"),
    "g1_guc_sold":            ("Sizer", "G66"),
    "g1_guc_retained":        ("Sizer", "G67"),
    # === Leverage deductions ===
    "declining_rents":        ("Sizer", "G84"),
    "declining_hp_appraisal": ("Sizer", "G85"),
    # === Loan Proceeds ===
    "initial_loan_amount":    ("Sizer", "G104"),
    "interest_reserves":      ("Sizer", "G106"),
    "financed_rehab":         ("Sizer", "G108"),
    "loan_interest_rate":     ("Sizer", "K121"),
    # === Exit Strategy DSCR ===
    "exit_dscr_rate":         ("Sizer", "T60"),
    # === Exit Strategy Agency ===
    "exit_agency_rate":       ("Sizer", "T48"),
    "exit_agency_advance":    ("Sizer", "T49"),
    "exit_agency_dscr":       ("Sizer", "T50"),
    # === HUD/Closing ===
    "lender_orig_pct":        ("Sizer", "X92"),
    "broker_orig_pct":        ("Sizer", "X94"),
    # === Liquidity ===
    "verified_liquidity":     ("Sizer", "X88"),
    "loan_id":                ("Sizer", "G124"),
}

GUC_INPUT_MAP = {
    # === Property Information ===
    "address":                ("Sizer", "G14"),
    "city":                   ("Sizer", "G15"),
    "state":                  ("Sizer", "G16"),
    "zip_code":               ("Sizer", "G17"),
    "num_units":              ("Sizer", "G19"),
    # === Loan Information ===
    "closing_date":           ("Sizer", "G24"),
    "closing_proceeds":       ("Sizer", "G25"),
    "loan_term":              ("Sizer", "G26"),
    # === Valuation & Rehab ===
    "purchase_price":         ("Sizer", "T14"),
    "purchase_date":          ("Sizer", "T15"),
    "appraisal_date":         ("Sizer", "T16"),
    "as_is_value":            ("Sizer", "T17"),
    "arv":                    ("Sizer", "T18"),
    "completed_rehab":        ("Sizer", "T19"),
    "rehab_budget":           ("Sizer", "T20"),
    "post_completion_sqft":   ("Sizer", "T21"),
    # === Borrower / Entity ===
    "entity_name":            ("Sizer", "G36"),
    "num_owners":             ("Sizer", "G37"),
    "guarantor_1_name":       ("Sizer", "G38"),
    "guarantor_1_fico":       ("Sizer", "G39"),
    "guarantor_1_credit_date":("Sizer", "G40"),
    "guarantor_1_ownership":  ("Sizer", "G41"),
    "guarantor_2_name":       ("Sizer", "G42"),
    "guarantor_2_fico":       ("Sizer", "G43"),
    "guarantor_2_credit_date":("Sizer", "G44"),
    "guarantor_2_ownership":  ("Sizer", "G45"),
    # === Experience ===
    "g1_construction_completed": ("Sizer", "T36"),  # Total GUC completed (sold + rented, summed in app.py)
    # === Leverage deductions ===
    "entitled_land":          ("Sizer", "G58"),
    "approved_permits":       ("Sizer", "G59"),
    "interest_reserves_flag": ("Sizer", "G60"),
    # === Loan Proceeds — user fills these in Excel themselves ===
    # (initial_loan_amount, interest_reserve_amt, financed_rehab removed per user request)
    # === Liquidity ===
    "verified_liquidity":     ("Sizer", "U63"),
    "loan_id":                ("Sizer", "G101"),
}

# Map loan types to their input maps
SIZER_MAPS = {
    "RTL": RTL_INPUT_MAP,
    "DSCR": DSCR_INPUT_MAP,
    "MF": MF_INPUT_MAP,
    "GUC": GUC_INPUT_MAP,
}

# Map loan types to template filenames
SIZER_TEMPLATES = {
    "RTL": "EV RTL Sizer_Jan 26 (5) copy.xlsx",
    "DSCR": "EV DSCR G Sizer_2.9.26 (4) copy.xlsx",
    "MF": "EV MF Sizer_Jan 25 (5) copy.xlsx",
    "GUC": "EV GUC Sizer_1.1 (3) copy.xlsx",
}


def fill_sizer(template_path: str, loan_type: str, inputs: dict) -> io.BytesIO:
    """
    Open the actual Eastview sizer template, fill in user-input cells,
    and return the workbook as a BytesIO stream.

    Args:
        template_path: Path to the .xlsx template file
        loan_type: One of "RTL", "DSCR", "MF", "GUC"
        inputs: Dict of { field_name: value } matching the input map keys

    Returns:
        BytesIO stream of the filled workbook
    """
    input_map = SIZER_MAPS.get(loan_type)
    if not input_map:
        raise ValueError(f"Unknown loan type: {loan_type}. Must be one of: {list(SIZER_MAPS.keys())}")

    # Load workbook preserving formulas, styles, and VBA
    wb = openpyxl.load_workbook(template_path)

    filled_count = 0
    for field_name, value in inputs.items():
        if field_name not in input_map:
            continue
        if value is None or value == "":
            continue

        sheet_name, cell_ref = input_map[field_name]

        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        ws[cell_ref] = value
        filled_count += 1

    # Write to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filled_count


def get_input_fields(loan_type: str) -> list[str]:
    """Return the list of fillable field names for a given loan type."""
    input_map = SIZER_MAPS.get(loan_type, {})
    return list(input_map.keys())


def get_template_path(assets_dir: str, loan_type: str) -> str:
    """Get the path to the template file for a given loan type."""
    filename = SIZER_TEMPLATES.get(loan_type)
    if not filename:
        raise ValueError(f"Unknown loan type: {loan_type}")
    return os.path.join(assets_dir, filename)
