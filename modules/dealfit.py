"""
modules/dealfit.py
Dealfit Rules Engine — evaluates a deal against multiple capital partner
guidelines (Colchis, Fidelis, Eastview) and returns eligibility, max leverage,
estimated pricing, and an AI-powered recommendation.

Workflow:
  1. User fills the A&S Capital Sizer Excel template
  2. Upload to Dealfit tab → AI (or direct cell reads) extracts deal parameters
  3. Each partner's evaluate() function checks eligibility & computes max leverage
  4. Claude AI generates a recommendation comparing all partners
"""

import io
import json
import os
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
import anthropic


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class DealParams:
    """All parameters extracted from the A&S Capital Sizer."""
    # Deal type
    deal_type: str = ""           # Fix & Flip, Bridge, Fix & Hold, Ground Up Construction
    transaction_type: str = ""    # Purchase, Refinance (Rate & Term), Refinance (Cash Out)
    loan_term: str = ""           # 6 Months, 12 Months, 13-18 Months, 19-24 Months
    deal_product: str = ""        # Light Rehab, Heavy Rehab, Bridge, Construction

    # Property
    address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    property_type: str = ""       # SFR, Townhome, Condo, PUD, 2-4 Unit, 5-10 MFR, etc.
    num_units: int = 1
    sqft: int = 0
    year_built: int = 0
    lot_size: int = 0
    condition: str = ""

    # Valuation
    purchase_price: float = 0
    as_is_value: float = 0
    arv: float = 0
    rehab_budget: float = 0

    # Loan request
    loan_amount: float = 0
    rehab_holdback: float = 0
    interest_reserve: float = 0

    # Borrower
    entity_name: str = ""
    num_guarantors: int = 1
    guarantor_1_name: str = ""
    guarantor_1_fico: int = 0
    guarantor_2_name: str = ""
    guarantor_2_fico: int = 0
    completed_projects: int = 0
    similar_experience: str = ""
    verified_liquidity: float = 0
    monthly_pitia: float = 0

    # ZHVI
    zhvi: float = 0
    value_zhvi_ratio: float = 0

    # Computed
    @property
    def total_loan(self) -> float:
        return self.loan_amount + self.rehab_holdback + self.interest_reserve

    @property
    def ltv(self) -> float:
        if self.as_is_value > 0:
            return self.loan_amount / self.as_is_value
        return 0

    @property
    def ltarv(self) -> float:
        if self.arv > 0:
            return self.total_loan / self.arv
        return 0

    @property
    def ltc(self) -> float:
        cost = self.purchase_price + self.rehab_budget
        if cost > 0:
            return self.total_loan / cost
        return 0

    @property
    def is_purchase(self) -> bool:
        return "purchase" in self.transaction_type.lower() if self.transaction_type else True

    @property
    def is_cashout_refi(self) -> bool:
        return "cash" in self.transaction_type.lower() if self.transaction_type else False

    @property
    def is_multifamily(self) -> bool:
        return self.num_units >= 5

    @property
    def primary_fico(self) -> int:
        return self.guarantor_1_fico


@dataclass
class PartnerResult:
    """Result of evaluating a deal against one capital partner."""
    partner_name: str
    eligible: bool = True
    ineligible_reasons: list = field(default_factory=list)

    # Max leverage the partner would allow
    max_ltv: float = 0
    max_ltarv: float = 0
    max_ltc: float = 0
    max_loan_amount: float = 0

    # Estimated rate
    estimated_rate: Optional[float] = None
    rate_notes: str = ""

    # Tier / classification
    experience_tier: str = ""
    fico_tier: str = ""

    # Warnings (deal is eligible but with conditions)
    warnings: list = field(default_factory=list)

    # Extra details
    details: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# A&S Sizer reader — extracts DealParams from the Excel template
# ---------------------------------------------------------------------------

def _safe_float(val, default=0) -> float:
    """Convert cell value to float, handling None and strings."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        cleaned = str(val).replace("$", "").replace(",", "").strip()
        return float(cleaned) if cleaned else default
    except (ValueError, TypeError):
        return default


def _safe_int(val, default=0) -> int:
    """Convert cell value to int."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return int(val)
    try:
        cleaned = str(val).replace(",", "").strip()
        return int(float(cleaned)) if cleaned else default
    except (ValueError, TypeError):
        return default


def _safe_str(val, default="") -> str:
    if val is None:
        return default
    return str(val).strip()


def read_sizer(file_bytes: bytes) -> DealParams:
    """Read the A&S Capital Sizer Excel template and return DealParams.

    Cell map (v2 — auto-calculating Eastview-quality layout):
        Sizer sheet:
          C5  = Deal Type          C6  = Transaction Type     C7  = Loan Term
          C8  = Deal Product
          C11 = Address (merged C11:D11)  C12 = City          E12 = State
          C13 = ZIP Code           C14 = Property Type        C15 = # Units
          C16 = Square Footage     E16 = Lot Size             C17 = Year Built (formula)
          C20 = Purchase Price     C21 = Purchase Date        C22 = As-Is Value
          C23 = ARV                C24 = Rehab Budget         C25 = Total Project Cost (formula)
          C28 = Initial Loan       C29 = Rehab Holdback       C30 = Interest Reserve
          C31 = Total Loan (formula)
          C34 = Borrowing Entity (merged C34:D34)   C35 = # Guarantors
          C38 = Guarantor 1 Name (merged C38:D38)   C39 = Guarantor 1 FICO
          C42 = Guarantor 2 Name (merged C42:D42)   C43 = Guarantor 2 FICO
          C46 = # Completed Projects  C47 = Similar Experience
          F20 = ZHVI (formula)        F21 = Value/ZHVI Ratio (formula)
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Sizer"]

    deal = DealParams()

    # Deal type
    deal.deal_type = _safe_str(ws["C5"].value)
    deal.transaction_type = _safe_str(ws["C6"].value)
    deal.loan_term = _safe_str(ws["C7"].value)
    deal.deal_product = _safe_str(ws["C8"].value)

    # Property
    deal.address = _safe_str(ws["C11"].value)
    deal.city = _safe_str(ws["C12"].value)
    deal.state = _safe_str(ws["E12"].value)
    deal.zip_code = _safe_str(ws["C13"].value)
    deal.property_type = _safe_str(ws["C14"].value)
    deal.num_units = _safe_int(ws["C15"].value, 1)
    deal.sqft = _safe_int(ws["C16"].value)
    deal.year_built = _safe_int(ws["C17"].value)
    deal.lot_size = _safe_int(ws["E16"].value)
    # condition field removed from sizer — no longer collected

    # Valuation
    deal.purchase_price = _safe_float(ws["C20"].value)
    deal.as_is_value = _safe_float(ws["C22"].value)
    deal.arv = _safe_float(ws["C23"].value)
    deal.rehab_budget = _safe_float(ws["C24"].value)

    # Loan
    deal.loan_amount = _safe_float(ws["C28"].value)
    deal.rehab_holdback = _safe_float(ws["C29"].value)
    deal.interest_reserve = _safe_float(ws["C30"].value)

    # Borrower
    deal.entity_name = _safe_str(ws["C34"].value)
    deal.num_guarantors = _safe_int(ws["C35"].value, 1)
    deal.guarantor_1_name = _safe_str(ws["C38"].value)
    deal.guarantor_1_fico = _safe_int(ws["C39"].value)
    deal.guarantor_2_name = _safe_str(ws["C42"].value)
    deal.guarantor_2_fico = _safe_int(ws["C43"].value)

    # Experience (liquidity & PITIA removed from sizer)
    deal.completed_projects = _safe_int(ws["C46"].value)
    deal.similar_experience = _safe_str(ws["C47"].value)

    # ZHVI (computed cells — data_only=True gives calculated values)
    deal.zhvi = _safe_float(ws["F20"].value)
    deal.value_zhvi_ratio = _safe_float(ws["F21"].value)

    # AIV fallback
    if deal.as_is_value == 0 and deal.purchase_price > 0:
        deal.as_is_value = deal.purchase_price

    wb.close()
    return deal


# ---------------------------------------------------------------------------
# COLCHIS CAPITAL — Rules Engine
# ---------------------------------------------------------------------------

def _colchis_experience_tier(projects: int) -> str:
    """Colchis experience tiers: 0-3, 4-7, 8+"""
    if projects >= 8:
        return "8+"
    elif projects >= 4:
        return "4-7"
    else:
        return "0-3"


def _colchis_fico_bucket(fico: int) -> str:
    if fico >= 740:
        return "740+"
    elif fico >= 700:
        return "700-739"
    elif fico >= 680:
        return "680-699"
    elif fico >= 660:
        return "660-679"
    else:
        return "<660"


# Colchis leverage grids: {deal_type: {fico_bucket: {exp_tier: (LTP_LTV, LTC, LTARV)}}}
# None means ineligible for that combination
COLCHIS_SF_LIGHT_REHAB = {
    "740+":   {"8+": (0.900, 0.925, 0.750), "4-7": (0.900, 0.925, 0.750), "0-3": (0.900, 0.900, 0.750)},
    "700-739":{"8+": (0.900, 0.925, 0.750), "4-7": (0.900, 0.925, 0.750), "0-3": (0.875, 0.900, 0.750)},
    "680-699":{"8+": (0.875, 0.900, 0.750), "4-7": (0.850, 0.875, 0.750), "0-3": (0.850, 0.850, 0.700)},
    "660-679":{"8+": None, "4-7": None, "0-3": None},
}

COLCHIS_SF_BRIDGE = {
    "740+":   {"8+": (0.750, None, None), "4-7": (0.750, None, None), "0-3": (0.750, None, None)},
    "700-739":{"8+": (0.750, None, None), "4-7": (0.750, None, None), "0-3": (0.700, None, None)},
    "680-699":{"8+": (0.700, None, None), "4-7": (0.700, None, None), "0-3": (0.650, None, None)},
}

COLCHIS_SF_HEAVY_REHAB = {
    "740+":   {"8+": (0.800, 0.850, 0.700), "4-7": (0.800, 0.850, 0.700), "0-3": None},
    "700-739":{"8+": (0.800, 0.850, 0.700), "4-7": (0.800, 0.850, 0.700), "0-3": None},
    "680-699":{"8+": (0.750, 0.825, 0.650), "4-7": (0.750, 0.800, 0.650), "0-3": None},
}

COLCHIS_SF_CONSTRUCTION = {
    "740+":   {"6+": (0.600, 0.900, 0.700), "4-5": (0.600, 0.850, 0.700), "0-3": None},
    "700-739":{"6+": (0.600, 0.900, 0.700), "4-5": (0.600, 0.850, 0.700), "0-3": None},
    "680-699":{"6+": (0.600, 0.850, 0.700), "4-5": (0.600, 0.825, 0.650), "0-3": None},
}

COLCHIS_SF_RT_REFI = {
    "740+":   {"8+": (0.750, None, None), "4-7": (0.750, None, None), "0-3": (0.750, None, None)},
    "700-739":{"8+": (0.750, None, None), "4-7": (0.750, None, None), "0-3": (0.700, None, None)},
    "680-699":{"8+": (0.700, None, None), "4-7": (0.700, None, None), "0-3": (0.650, None, None)},
}

COLCHIS_SF_CO_REFI = {
    "740+":   {"8+": (0.700, None, None), "4-7": (0.700, None, None), "0-3": (0.700, None, None)},
    "700-739":{"8+": (0.700, None, None), "4-7": (0.700, None, None), "0-3": (0.650, None, None)},
    "680-699":{"8+": (0.650, None, None), "4-7": (0.650, None, None), "0-3": (0.600, None, None)},
}

# Colchis MF grids (5-10 units)
COLCHIS_MF_LIGHT_REHAB = {
    "740+":   {"8+": (0.800, 0.850, 0.650), "4-7": (0.800, 0.850, 0.650), "0-3": None},
    "700-739":{"8+": (0.800, 0.850, 0.650), "4-7": (0.800, 0.800, 0.650), "0-3": None},
    "680-699":{"8+": (0.750, 0.800, 0.600), "4-7": (0.750, 0.800, 0.600), "0-3": None},
}

COLCHIS_MF_BRIDGE = {
    "740+":   {"8+": (0.700, None, 0.650), "4-7": (0.700, None, 0.650), "0-3": None},
    "700-739":{"8+": (0.700, None, 0.600), "4-7": (0.700, None, 0.600), "0-3": None},
    "680-699":{"8+": (0.650, None, None),   "4-7": (0.600, None, None),   "0-3": None},
}

# Colchis pricing grids: {product: {fico_bucket: {ltc_bucket: rate}}}
COLCHIS_PRICING_BRIDGE = {
    "740+":   {"<=70%": 0.07750, "<=75%": 0.07750},
    "700-739":{"<=70%": 0.07750, "<=75%": 0.07750},
    "680-699":{"<=70%": 0.07875, "<=75%": None},
}

COLCHIS_PRICING_LIGHT_REHAB = {
    "740+":   {"<=70%": 0.07750, "<=75%": 0.07750, "<=80%": 0.07750, "<=85%": 0.07875, "<=90%": 0.08000, "<=95%": 0.08250},
    "700-739":{"<=70%": 0.07750, "<=75%": 0.07750, "<=80%": 0.07875, "<=85%": 0.08000, "<=90%": 0.08125, "<=95%": 0.08375},
    "680-699":{"<=70%": 0.07875, "<=75%": 0.08000, "<=80%": 0.08125, "<=85%": 0.08250, "<=90%": 0.08375, "<=95%": None},
}

COLCHIS_PRICING_HEAVY_REHAB = {
    "740+":   {"<=70%": 0.08375, "<=75%": 0.08375, "<=80%": 0.08500, "<=85%": 0.08625},
    "700-739":{"<=70%": 0.08375, "<=75%": 0.08500, "<=80%": 0.08625, "<=85%": 0.08750},
    "680-699":{"<=70%": 0.08625, "<=75%": 0.08750, "<=80%": 0.08875, "<=85%": 0.09000},
}

COLCHIS_PRICING_CONSTRUCTION = {
    "740+":   {"<=70%": 0.08375, "<=75%": 0.08375, "<=80%": 0.08500, "<=85%": 0.08625, "<=90%": 0.08875},
    "700-739":{"<=70%": 0.08375, "<=75%": 0.08500, "<=80%": 0.08625, "<=85%": 0.08750, "<=90%": 0.09000},
    "680-699":{"<=70%": 0.08625, "<=75%": 0.08750, "<=80%": 0.08875, "<=85%": 0.09000, "<=90%": None},
}


def _colchis_get_ltc_bucket(ltc: float) -> str:
    """Map LTC ratio to Colchis pricing bucket."""
    if ltc <= 0.70:
        return "<=70%"
    elif ltc <= 0.75:
        return "<=75%"
    elif ltc <= 0.80:
        return "<=80%"
    elif ltc <= 0.85:
        return "<=85%"
    elif ltc <= 0.90:
        return "<=90%"
    else:
        return "<=95%"


def _colchis_base_rate(deal: DealParams, fico_bucket: str) -> Optional[float]:
    """Look up Colchis base rate from pricing grid."""
    ltc = deal.ltc
    ltc_bucket = _colchis_get_ltc_bucket(ltc)

    dt = deal.deal_type.lower()
    if "ground up" in dt or "construction" in dt or "guc" in dt:
        grid = COLCHIS_PRICING_CONSTRUCTION
    elif "bridge" in dt:
        grid = COLCHIS_PRICING_BRIDGE
    elif deal.rehab_budget > 0:
        # Determine light vs heavy rehab
        project_cost = deal.purchase_price + deal.rehab_budget
        rehab_pct = deal.rehab_budget / project_cost if project_cost > 0 else 0
        # Heavy if >= 50% of project costs, or budget > $250K (>$300K in CA)
        ca_threshold = 300000 if deal.state == "CA" else 250000
        if rehab_pct >= 0.50 or deal.rehab_budget > ca_threshold:
            grid = COLCHIS_PRICING_HEAVY_REHAB
        else:
            grid = COLCHIS_PRICING_LIGHT_REHAB
    else:
        grid = COLCHIS_PRICING_BRIDGE  # No rehab = bridge

    fico_rates = grid.get(fico_bucket, {})
    rate = fico_rates.get(ltc_bucket)
    return rate


def _colchis_rate_adjustments(deal: DealParams, exp_tier: str) -> float:
    """Calculate Colchis rate adjustments."""
    adj = 0.0

    # Experience adjustment
    if deal.completed_projects >= 8:
        adj -= 0.00250  # Tier 1 discount
    elif deal.completed_projects <= 3:
        adj += 0.00250  # Tier 3 surcharge

    # Loan term
    term = deal.loan_term.lower()
    if "19" in term or "24" in term:
        adj += 0.00125

    # Cash-out refinance
    if deal.is_cashout_refi:
        adj += 0.00250

    # Loan amount
    if deal.total_loan > 3000000:
        adj += 0.00125

    # Geographic adjustments
    if deal.state in ("NY", "NJ", "CT"):
        adj += 0.00250
    elif deal.state == "CA":
        adj -= 0.00125

    # ZHVI adjustments
    if deal.value_zhvi_ratio > 3.0:
        adj += 0.00375
    elif deal.value_zhvi_ratio > 2.0:
        adj += 0.00125

    return adj


def evaluate_colchis(deal: DealParams) -> PartnerResult:
    """Evaluate a deal against Colchis Capital guidelines."""
    result = PartnerResult(partner_name="Colchis Capital")

    fico = deal.primary_fico
    fico_bucket = _colchis_fico_bucket(fico)
    exp_tier = _colchis_experience_tier(deal.completed_projects)
    result.fico_tier = fico_bucket
    result.experience_tier = exp_tier

    # --- Hard eligibility checks ---

    # Min FICO
    if fico < 680:
        result.eligible = False
        result.ineligible_reasons.append(f"FICO {fico} below Colchis minimum of 680")

    # Geographic exclusions
    if deal.state == "IL":
        result.eligible = False
        result.ineligible_reasons.append("Illinois is excluded by Colchis")
    if deal.city.lower() == "newark" and deal.state == "NJ":
        result.eligible = False
        result.ineligible_reasons.append("Newark, NJ is excluded by Colchis")

    # Loan amount limits
    if deal.total_loan < 100000:
        result.eligible = False
        result.ineligible_reasons.append(f"Total loan ${deal.total_loan:,.0f} below Colchis minimum of $100,000")
    if deal.total_loan > 3500000:
        result.eligible = False
        result.ineligible_reasons.append(f"Total loan ${deal.total_loan:,.0f} exceeds Colchis maximum of $3,500,000")

    # Property type
    if deal.num_units > 20:
        result.eligible = False
        result.ineligible_reasons.append(f"{deal.num_units} units exceeds Colchis max of 20 units")

    if not result.eligible:
        return result

    # --- Determine leverage grid ---
    dt = deal.deal_type.lower()
    is_mf = deal.is_multifamily

    if is_mf:
        if "bridge" in dt:
            grid = COLCHIS_MF_BRIDGE
        elif "ground up" in dt or "construction" in dt or "guc" in dt:
            result.eligible = False
            result.ineligible_reasons.append("Colchis does not offer MF Heavy Rehab / Construction")
            return result
        else:
            grid = COLCHIS_MF_LIGHT_REHAB
    else:
        if "ground up" in dt or "construction" in dt or "guc" in dt:
            grid = COLCHIS_SF_CONSTRUCTION
            # Construction uses different experience tiers
            if deal.completed_projects >= 6:
                exp_tier = "6+"
            elif deal.completed_projects >= 4:
                exp_tier = "4-5"
            else:
                exp_tier = "0-3"
        elif "bridge" in dt:
            if deal.is_cashout_refi:
                grid = COLCHIS_SF_CO_REFI
            elif not deal.is_purchase:
                grid = COLCHIS_SF_RT_REFI
            else:
                grid = COLCHIS_SF_BRIDGE
        elif deal.rehab_budget > 0:
            # Determine light vs heavy
            project_cost = deal.purchase_price + deal.rehab_budget
            rehab_pct = deal.rehab_budget / project_cost if project_cost > 0 else 0
            ca_threshold = 300000 if deal.state == "CA" else 250000
            if rehab_pct >= 0.50 or deal.rehab_budget > ca_threshold:
                grid = COLCHIS_SF_HEAVY_REHAB
            else:
                grid = COLCHIS_SF_LIGHT_REHAB
        else:
            if deal.is_cashout_refi:
                grid = COLCHIS_SF_CO_REFI
            elif not deal.is_purchase:
                grid = COLCHIS_SF_RT_REFI
            else:
                grid = COLCHIS_SF_BRIDGE

    # Look up leverage limits
    fico_row = grid.get(fico_bucket)
    if not fico_row:
        result.eligible = False
        result.ineligible_reasons.append(f"FICO {fico} ({fico_bucket}) not eligible at Colchis for this deal type")
        return result

    leverage = fico_row.get(exp_tier)
    if leverage is None:
        result.eligible = False
        result.ineligible_reasons.append(
            f"Experience tier {exp_tier} ({deal.completed_projects} projects) not eligible at Colchis for FICO {fico_bucket}"
        )
        return result

    max_ltv, max_ltc, max_ltarv = leverage
    result.max_ltv = max_ltv or 0
    result.max_ltc = max_ltc or 0
    result.max_ltarv = max_ltarv or 0
    result.max_loan_amount = 3500000

    # ZHVI adjustments to LTV
    if deal.value_zhvi_ratio > 3.0:
        result.max_ltv = max(0, result.max_ltv - 0.10)
        result.warnings.append("High-value property (>300% ZHVI): LTV reduced by 10%")
    elif deal.value_zhvi_ratio > 2.0:
        result.max_ltv = max(0, result.max_ltv - 0.05)
        result.warnings.append("High-value property (>200% ZHVI): LTV reduced by 5%")

    # Check if deal's leverage fits within limits
    if result.max_ltv > 0 and deal.ltv > result.max_ltv:
        result.warnings.append(
            f"Deal LTV {deal.ltv:.1%} exceeds max {result.max_ltv:.1%} — reduce loan amount"
        )
    if result.max_ltarv > 0 and deal.ltarv > result.max_ltarv:
        result.warnings.append(
            f"Deal LTARV {deal.ltarv:.1%} exceeds max {result.max_ltarv:.1%}"
        )
    if result.max_ltc > 0 and deal.ltc > result.max_ltc:
        result.warnings.append(
            f"Deal LTC {deal.ltc:.1%} exceeds max {result.max_ltc:.1%}"
        )

    # Construction: 90% LTC requires budget < $500K
    if "construction" in dt or "guc" in dt or "ground up" in dt:
        if result.max_ltc >= 0.90 and deal.rehab_budget >= 500000:
            result.max_ltc = 0.85
            result.warnings.append("LTC capped at 85% (construction budget >= $500K)")

    # --- Pricing ---
    base_rate = _colchis_base_rate(deal, fico_bucket)
    if base_rate is not None:
        adjustments = _colchis_rate_adjustments(deal, exp_tier)
        result.estimated_rate = base_rate + adjustments

        # MFR adder
        if is_mf:
            mfr_add = 0.00375 if deal.ltc <= 0.70 else 0.00500
            result.estimated_rate += mfr_add
            result.rate_notes += f"Includes MFR adder (+{mfr_add:.3%}). "

        # Min rate floors
        min_rates = {
            "bridge": 0.07750,
            "light": 0.07750,
            "heavy": 0.08375,
            "construction": 0.08375,
        }
        for key, floor in min_rates.items():
            if key in dt.lower() or (key == "construction" and ("guc" in dt.lower() or "ground up" in dt.lower())):
                result.estimated_rate = max(result.estimated_rate, floor)
                break
        else:
            result.estimated_rate = max(result.estimated_rate, 0.07750)

        if adjustments != 0:
            result.rate_notes += f"Adjustments: {adjustments:+.3%}"
    else:
        result.rate_notes = "Rate not available for this LTC/FICO combination"

    return result


# ---------------------------------------------------------------------------
# FIDELIS INVESTORS — Rules Engine
# ---------------------------------------------------------------------------

def _fidelis_experience_tier(projects: int) -> str:
    """Fidelis tiers: Tier 1 (3+), Tier 2 (1-2), Tier 3 (0)"""
    if projects >= 3:
        return "Tier 1"
    elif projects >= 1:
        return "Tier 2"
    else:
        return "Tier 3"


# Fidelis leverage grids: {state_group: {deal_category: {tier: (max_loan, min_fico, max_acq_ltv, max_arltv, max_ltc)}}}
# None means ineligible

FIDELIS_NATIONAL = {
    "purchase_ff": {
        "Tier 1": (3000000, 600, 0.900, 0.750, 0.925),
        "Tier 2": (2500000, 660, 0.900, 0.700, 0.925),
        "Tier 3": (950000,  680, 0.800, 0.650, 0.850),
    },
    "purchase_nc": {
        "Tier 1": (2500000, 680, 0.700, 0.700, 0.850),
        "Tier 2": (2500000, 700, 0.700, 0.650, 0.850),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.800),
    },
    "purchase_bridge": {
        "Tier 1": (2500000, 600, 0.750, 0.750, 0.750),
        "Tier 2": (2500000, 660, 0.700, 0.700, 0.700),
        "Tier 3": (950000,  680, 0.700, 0.700, 0.700),
    },
    "refi_ff": {
        "Tier 1": (2500000, 640, 0.800, 0.700, 0.850),
        "Tier 2": (2500000, 660, 0.750, 0.650, 0.800),
        "Tier 3": None,
    },
    "refi_nc": {
        "Tier 1": (2500000, 680, 0.700, 0.700, 0.850),
        "Tier 2": (2500000, 700, 0.700, 0.650, 0.850),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.800),
    },
    "refi_bridge": {
        "Tier 1": (1500000, 680, 0.700, 0.700, 0.700),
        "Tier 2": (1500000, 700, 0.700, 0.700, 0.700),
        "Tier 3": (950000,  740, 0.650, 0.650, 0.650),
    },
}

FIDELIS_FLORIDA = {
    "purchase_ff": {
        "Tier 1": (1500000, 700, 0.850, 0.700, 0.875),
        "Tier 2": (1500000, 730, 0.850, 0.700, 0.850),
        "Tier 3": (950000,  740, 0.800, 0.650, 0.850),
    },
    "purchase_nc": {
        "Tier 1": (1500000, 700, 0.700, 0.700, 0.800),
        "Tier 2": (1500000, 730, 0.700, 0.650, 0.800),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.800),
    },
    "purchase_bridge": {
        "Tier 1": (1500000, 700, 0.750, 0.750, 0.750),
        "Tier 2": (1500000, 730, 0.700, 0.700, 0.700),
        "Tier 3": (950000,  740, 0.700, 0.700, 0.700),
    },
    "refi_ff": {
        "Tier 1": None,  # FL refi F&F ineligible
        "Tier 2": None,
        "Tier 3": None,
    },
    "refi_nc": {
        "Tier 1": (1500000, 700, 0.700, 0.700, 0.800),
        "Tier 2": (1500000, 730, 0.700, 0.650, 0.800),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.800),
    },
    "refi_bridge": {
        "Tier 1": (1500000, 700, 0.650, 0.650, 0.650),
        "Tier 2": (1500000, 730, 0.650, 0.650, 0.650),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.600),
    },
}

FIDELIS_CA_NY = {
    "purchase_ff": {
        "Tier 1": (3500000, 640, 0.900, 0.750, 0.925),
        "Tier 2": (3500000, 660, 0.900, 0.700, 0.925),
        "Tier 3": (950000,  680, 0.800, 0.700, 0.850),
    },
    "purchase_nc": {
        "Tier 1": (3500000, 680, 0.700, 0.700, 0.850),
        "Tier 2": (3500000, 700, 0.700, 0.650, 0.850),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.800),
    },
    "purchase_bridge": {
        "Tier 1": (3500000, 600, 0.750, 0.750, 0.750),
        "Tier 2": (3500000, 660, 0.700, 0.700, 0.700),
        "Tier 3": (950000,  680, 0.700, 0.700, 0.700),
    },
    "refi_ff": {
        "Tier 1": None,  # CA/NY refi F&F ineligible
        "Tier 2": None,
        "Tier 3": None,
    },
    "refi_nc": {
        "Tier 1": (3500000, 680, 0.700, 0.700, 0.850),
        "Tier 2": (3500000, 700, 0.700, 0.650, 0.850),
        "Tier 3": (950000,  740, 0.600, 0.600, 0.800),
    },
    "refi_bridge": {
        "Tier 1": (2500000, 680, 0.700, 0.700, 0.700),
        "Tier 2": (2500000, 700, 0.700, 0.700, 0.700),
        "Tier 3": (950000,  740, 0.650, 0.650, 0.650),
    },
}

FIDELIS_MF = {
    "purchase_ff": {
        "Tier 1": (8500000, 650, 0.825, 0.750, 0.850),
        "Tier 2": (4500000, 660, 0.775, 0.700, 0.800),
        "Tier 3": None,
    },
    "purchase_nc": {
        "Tier 1": (8500000, 650, 0.700, 0.700, 0.850),
        "Tier 2": (4500000, 660, 0.650, 0.650, 0.800),
        "Tier 3": None,
    },
    "purchase_bridge": {
        "Tier 1": (8500000, 650, 0.750, 0.750, 0.750),
        "Tier 2": (4500000, 660, 0.700, 0.700, 0.700),
        "Tier 3": None,
    },
    "refi_ff": {
        "Tier 1": (8500000, 650, 0.800, 0.700, 0.825),
        "Tier 2": (4500000, 660, 0.750, 0.650, 0.775),
        "Tier 3": None,
    },
    "refi_nc": {
        "Tier 1": (8500000, 650, 0.700, 0.700, 0.850),
        "Tier 2": (4500000, 660, 0.650, 0.650, 0.800),
        "Tier 3": None,
    },
    "refi_bridge": {
        "Tier 1": (8500000, 650, 0.700, 0.700, 0.700),
        "Tier 2": (4500000, 660, 0.650, 0.650, 0.650),
        "Tier 3": None,
    },
}


def _fidelis_deal_category(deal: DealParams) -> str:
    """Map deal type + transaction type to Fidelis grid category key."""
    dt = deal.deal_type.lower()
    is_nc = "ground up" in dt or "construction" in dt or "guc" in dt
    is_bridge = "bridge" in dt
    # fix & flip and fix & hold use same grid
    is_ff = "flip" in dt or "hold" in dt or (not is_nc and not is_bridge)

    if deal.is_purchase:
        if is_nc:
            return "purchase_nc"
        elif is_bridge:
            return "purchase_bridge"
        else:
            return "purchase_ff"
    else:
        if is_nc:
            return "refi_nc"
        elif is_bridge:
            return "refi_bridge"
        else:
            return "refi_ff"


def evaluate_fidelis(deal: DealParams) -> PartnerResult:
    """Evaluate a deal against Fidelis Investors guidelines."""
    result = PartnerResult(partner_name="Fidelis Investors")

    fico = deal.primary_fico
    tier = _fidelis_experience_tier(deal.completed_projects)
    result.experience_tier = tier
    result.fico_tier = f"FICO {fico}"

    # --- Hard eligibility checks ---

    # Min FICO by tier
    min_ficos = {"Tier 1": 600, "Tier 2": 660, "Tier 3": 680}
    min_fico = min_ficos.get(tier, 680)
    if fico < min_fico:
        result.eligible = False
        result.ineligible_reasons.append(
            f"FICO {fico} below Fidelis minimum of {min_fico} for {tier}"
        )

    # Loan amount
    if deal.total_loan < 75000:
        result.eligible = False
        result.ineligible_reasons.append(f"Total loan ${deal.total_loan:,.0f} below Fidelis minimum of $75,000")

    # Ineligible property types
    pt = deal.property_type.lower()
    if "mobile" in pt or "coop" in pt or "cooperative" in pt:
        result.eligible = False
        result.ineligible_reasons.append(f"Property type '{deal.property_type}' ineligible at Fidelis")

    # MF requirements
    if deal.is_multifamily:
        if deal.num_units > 50:
            result.eligible = False
            result.ineligible_reasons.append(f"{deal.num_units} units exceeds Fidelis max of 50")
        if deal.completed_projects < 1:
            result.eligible = False
            result.ineligible_reasons.append("Fidelis MF requires at least 1 completed project (Tier 3 ineligible)")
        # Min ARV per unit
        if deal.arv > 0 and deal.num_units > 0:
            arv_per_unit = deal.arv / deal.num_units
            if arv_per_unit < 100000:
                result.eligible = False
                result.ineligible_reasons.append(
                    f"ARV/unit ${arv_per_unit:,.0f} below Fidelis MF minimum of $100,000/unit"
                )

    # FL/CA/NY refi restrictions
    if deal.state in ("FL", "CA", "NY") and not deal.is_purchase:
        dt_lower = deal.deal_type.lower()
        if "flip" in dt_lower or "hold" in dt_lower:
            result.eligible = False
            result.ineligible_reasons.append(
                f"Refinance of Fix & Flip/Hold in {deal.state} is ineligible at Fidelis"
            )

    if not result.eligible:
        return result

    # --- Select correct grid ---
    category = _fidelis_deal_category(deal)

    if deal.is_multifamily:
        grid = FIDELIS_MF
    elif deal.state == "FL":
        grid = FIDELIS_FLORIDA
    elif deal.state in ("CA", "NY"):
        grid = FIDELIS_CA_NY
    else:
        grid = FIDELIS_NATIONAL

    category_data = grid.get(category, {})
    tier_data = category_data.get(tier)

    if tier_data is None:
        result.eligible = False
        result.ineligible_reasons.append(
            f"{tier} ({deal.completed_projects} projects) not eligible at Fidelis for {category.replace('_', ' ')}"
        )
        return result

    max_loan, min_fico_req, max_ltv, max_arltv, max_ltc = tier_data

    # Check FICO against tier-specific minimum
    if fico < min_fico_req:
        result.eligible = False
        result.ineligible_reasons.append(
            f"FICO {fico} below Fidelis minimum of {min_fico_req} for {tier} {category.replace('_', ' ')}"
        )
        return result

    result.max_ltv = max_ltv
    result.max_ltarv = max_arltv
    result.max_ltc = max_ltc
    result.max_loan_amount = max_loan

    # Cash-out refi MF: -10% leverage
    if deal.is_cashout_refi and deal.is_multifamily:
        result.max_ltv = max(0, result.max_ltv - 0.10)
        result.max_ltarv = max(0, result.max_ltarv - 0.10)
        result.max_ltc = max(0, result.max_ltc - 0.10)
        result.warnings.append("Cash-out MF refinance: leverage reduced by 10%")

    # Check loan amount limit
    if deal.total_loan > max_loan:
        result.warnings.append(
            f"Total loan ${deal.total_loan:,.0f} exceeds Fidelis max ${max_loan:,.0f} for {tier}"
        )

    # Check leverage
    if result.max_ltv > 0 and deal.ltv > result.max_ltv:
        result.warnings.append(
            f"Deal LTV {deal.ltv:.1%} exceeds max {result.max_ltv:.1%}"
        )
    if result.max_ltarv > 0 and deal.ltarv > result.max_ltarv:
        result.warnings.append(
            f"Deal LTARV {deal.ltarv:.1%} exceeds max {result.max_ltarv:.1%}"
        )
    if result.max_ltc > 0 and deal.ltc > result.max_ltc:
        result.warnings.append(
            f"Deal LTC {deal.ltc:.1%} exceeds max {result.max_ltc:.1%}"
        )

    # Liquidity check
    if deal.total_loan >= 1000000:
        required_months = 4
    else:
        required_months = 2
    if deal.monthly_pitia > 0:
        months_reserves = deal.verified_liquidity / deal.monthly_pitia
        if months_reserves < required_months:
            result.warnings.append(
                f"Liquidity: {months_reserves:.1f} months reserves vs {required_months} required"
            )

    # Pricing placeholder — will be populated when rate sheet is provided
    result.estimated_rate = None
    result.rate_notes = "Fidelis rate sheet pending — contact for pricing"

    return result


# ---------------------------------------------------------------------------
# EASTVIEW CAPITAL — Rules Engine (RTL v4.2, GUC v1.1, MF v1.0)
# ---------------------------------------------------------------------------

def _eastview_rtl_classification(fico: int, experience: int) -> str:
    """Eastview RTL borrower classification: A+, A, B, C based on score."""
    # Credit Decision Score
    if fico >= 700:
        credit_score = 3
    elif fico >= 680:
        credit_score = 1
    else:
        credit_score = 0  # <680 or Foreign National

    # Experience Score
    if experience >= 10:
        exp_score = 7
    elif experience >= 3:
        exp_score = 5
    else:
        exp_score = 1  # 0-2

    total = credit_score + exp_score
    if total >= 7:
        return "A+"
    elif total >= 5:
        return "A"
    elif total >= 2:
        return "B"
    else:
        return "C"


def _eastview_guc_classification(experience: int, fico: int) -> str:
    """Eastview GUC borrower classification: A+, A, B."""
    if experience >= 4:
        return "A+"
    elif experience >= 2:
        return "A"
    elif experience >= 1 and fico >= 700:
        return "B"
    else:
        return None  # Ineligible


def _eastview_mf_classification(fico: int, experience: int) -> str:
    """Eastview MF borrower classification: A+, A (score-based)."""
    # Credit Decision Score (MF uses different thresholds)
    if fico >= 750:
        credit_score = 3
    elif fico >= 700:
        credit_score = 2
    else:
        credit_score = 0

    # Experience Score (MF uses different tiers)
    if experience >= 20:
        exp_score = 7
    elif experience >= 10:
        exp_score = 5
    elif experience >= 3:
        exp_score = 3
    else:
        exp_score = 1

    total = credit_score + exp_score
    if total >= 7:
        return "A+"
    elif total >= 5:
        return "A"
    else:
        return None  # Only A+ and A eligible for MF


# Eastview RTL leverage: {product: {purpose: {class: (LTV, LTC, LTARV)}}}
EV_RTL_FF = {
    "purchase":   {"A+": (0.900, 0.900, 0.750), "A": (0.850, 0.850, 0.700), "B": (0.825, 0.825, 0.650), "C": (0.750, 0.750, 0.600)},
    "refi_rt":    {"A+": (0.750, None, 0.650),   "A": (0.725, None, 0.650),   "B": (0.700, None, 0.600),   "C": (0.600, None, 0.550)},
    "refi_co":    {"A+": (0.700, None, 0.650),   "A": (0.675, None, 0.650),   "B": (0.650, None, 0.600),   "C": None},
}

EV_RTL_BRIDGE = {
    "purchase":   {"A+": (0.825, 0.825, None), "A": (0.800, 0.800, None), "B": (0.800, 0.800, None), "C": (0.750, 0.750, None)},
    "refi_rt":    {"A+": (0.775, None, None),   "A": (0.750, None, None),   "B": (0.750, None, None),   "C": (0.650, None, None)},
    "refi_co":    {"A+": (0.725, None, None),   "A": (0.700, None, None),   "B": (0.700, None, None),   "C": (0.600, None, None)},
}

EV_RTL_BRIDGE_PLUS = {
    "purchase":   {"A+": (0.775, 0.775, None), "A": (0.750, 0.750, None), "B": (0.750, 0.750, None), "C": (0.700, 0.700, None)},
    "refi_rt":    {"A+": (0.725, None, None),   "A": (0.700, None, None),   "B": (0.700, None, None),   "C": (0.600, None, None)},
    "refi_co":    {"A+": (0.675, None, None),   "A": (0.650, None, None),   "B": (0.650, None, None),   "C": (0.550, None, None)},
}

# Eastview GUC leverage: {purpose: {class: (As-Is LTV, LTC, TLTC, LTARV)}}
EV_GUC = {
    "purchase":   {"A+": (0.750, 0.750, 0.900, 0.700), "A": (0.700, 0.700, 0.850, 0.700), "B": (0.650, 0.650, 0.800, 0.650)},
    "refi_rt":    {"A+": (0.700, None, None, 0.700),     "A": (0.650, None, None, 0.700),     "B": (0.600, None, None, 0.650)},
    "refi_co":    {"A+": (0.650, None, None, 0.650),     "A": (0.600, None, None, 0.650),     "B": (0.550, None, None, 0.600)},
}

# Eastview MF leverage: {product: {purpose: {class: (LTV, LTC, LTARV)}}}
EV_MF_CAPEX = {
    "purchase":   {"A+": (0.800, 0.800, 0.700), "A": (0.800, 0.800, 0.650)},
    "refi_rt":    {"A+": (0.750, None, 0.650),   "A": (0.750, None, 0.600)},
    "refi_co":    {"A+": (0.650, None, 0.550),   "A": (0.650, None, 0.500)},
}

EV_MF_BRIDGE = {
    "purchase":   {"A+": (0.750, 0.750, None), "A": (0.725, 0.725, None)},
    "refi_rt":    {"A+": (0.700, None, None),   "A": (0.675, None, None)},
    "refi_co":    {"A+": (0.550, None, None),   "A": (0.500, None, None)},
}

EV_NOT_PERMITTED = {"ND", "SD", "HI", "AK"}


def evaluate_eastview(deal: DealParams) -> PartnerResult:
    """Evaluate a deal against Eastview Capital guidelines (RTL v4.2, GUC v1.1, MF v1.0)."""
    result = PartnerResult(partner_name="Eastview Capital")

    fico = deal.primary_fico
    exp = deal.completed_projects
    dt = deal.deal_type.lower()
    is_guc = "ground up" in dt or "construction" in dt or "guc" in dt
    is_bridge = "bridge" in dt
    is_mf = deal.is_multifamily

    # --- Hard eligibility checks ---

    # Min FICO
    if is_mf:
        min_fico = 680
    else:
        min_fico = 660
    if fico < min_fico and fico > 0:
        result.eligible = False
        result.ineligible_reasons.append(f"FICO {fico} below Eastview minimum of {min_fico}")

    # Not Permitted States
    if deal.state in EV_NOT_PERMITTED:
        result.eligible = False
        result.ineligible_reasons.append(f"{deal.state} is not permitted by Eastview")

    # Loan amount limits
    if is_mf:
        min_loan = 500000
        max_loan = 3500000
        if deal.num_units >= 9:
            min_loan = 1000000
    elif is_guc:
        min_loan = 150000
        max_loan = 1500000
    else:
        min_loan = 100000
        max_loan = 3000000

    if deal.total_loan < min_loan:
        result.eligible = False
        result.ineligible_reasons.append(f"Total loan ${deal.total_loan:,.0f} below Eastview min ${min_loan:,.0f}")
    if deal.total_loan > max_loan:
        result.warnings.append(f"Total loan ${deal.total_loan:,.0f} exceeds standard max ${max_loan:,.0f} (exception may apply)")

    # MF: max $400K per unit
    if is_mf and deal.num_units > 0:
        loan_per_unit = deal.total_loan / deal.num_units
        if loan_per_unit > 400000:
            result.eligible = False
            result.ineligible_reasons.append(f"Loan/unit ${loan_per_unit:,.0f} exceeds Eastview max $400,000/unit")
        # Min value per unit
        if deal.arv > 0:
            val_per_unit = deal.arv / deal.num_units
            if val_per_unit < 100000:
                result.warnings.append(f"Value/unit ${val_per_unit:,.0f} below standard min $100K (exception $50-99K)")

    # Property type (RTL/GUC: 1-4 units only; MF: 5+ units)
    if not is_mf and deal.num_units > 4:
        result.eligible = False
        result.ineligible_reasons.append(f"{deal.num_units} units — use Eastview MF product for 5+ units")

    if not result.eligible:
        return result

    # --- Classification and leverage lookup ---
    purpose = "purchase" if deal.is_purchase else ("refi_co" if deal.is_cashout_refi else "refi_rt")

    if is_mf:
        bclass = _eastview_mf_classification(fico, exp)
        if bclass is None:
            result.eligible = False
            result.ineligible_reasons.append(
                f"Eastview MF requires A+ or A classification (FICO {fico}, {exp} projects = score too low)"
            )
            return result

        result.experience_tier = bclass
        result.fico_tier = f"FICO {fico}"

        # CAPEX (has rehab) vs Bridge (no rehab)
        if deal.rehab_budget > 0:
            grid = EV_MF_CAPEX
        else:
            grid = EV_MF_BRIDGE

        purpose_data = grid.get(purpose, {})
        leverage = purpose_data.get(bclass)
        if leverage is None:
            result.eligible = False
            result.ineligible_reasons.append(f"Eastview MF: {bclass} not eligible for {purpose}")
            return result

        max_ltv, max_ltc, max_ltarv = leverage
        result.max_ltv = max_ltv or 0
        result.max_ltc = max_ltc or 0
        result.max_ltarv = max_ltarv or 0
        result.max_loan_amount = max_loan

        # MF leverage reductions
        if deal.rehab_budget > 0 and deal.purchase_price > 0:
            if deal.rehab_budget > deal.purchase_price:  # Heavy rehab
                result.max_ltv = max(0, result.max_ltv - 0.05)
                if result.max_ltc:
                    result.max_ltc = max(0, result.max_ltc - 0.05)
                if result.max_ltarv:
                    result.max_ltarv = max(0, result.max_ltarv - 0.05)
                result.warnings.append("Heavy rehab: -5% leverage reduction")

    elif is_guc:
        bclass = _eastview_guc_classification(exp, fico)
        if bclass is None:
            result.eligible = False
            result.ineligible_reasons.append(
                f"Eastview GUC requires min 1 GUC project (you have {exp}) and FICO >= 700 for Class B"
            )
            return result

        result.experience_tier = bclass
        result.fico_tier = f"FICO {fico}"

        purpose_data = EV_GUC.get(purpose, {})
        leverage = purpose_data.get(bclass)
        if leverage is None:
            result.eligible = False
            result.ineligible_reasons.append(f"Eastview GUC: {bclass} not eligible for {purpose}")
            return result

        max_ltv, max_ltc, max_tltc, max_ltarv = leverage
        result.max_ltv = max_ltv or 0
        result.max_ltc = max_tltc or max_ltc or 0  # Use TLTC if available
        result.max_ltarv = max_ltarv or 0
        result.max_loan_amount = max_loan

        # FICO 660-679 cap at 65%
        if 660 <= fico <= 679:
            result.max_ltv = min(result.max_ltv, 0.65)
            if result.max_ltc:
                result.max_ltc = min(result.max_ltc, 0.65)
            result.warnings.append("FICO 660-679: LTV/LTC capped at 65%")

        # ZHVI / HPA reductions
        if deal.value_zhvi_ratio > 3.0:
            result.max_ltv = max(0, result.max_ltv - 0.10)
            result.warnings.append("ZHVI >300%: -10% leverage reduction")
        elif deal.value_zhvi_ratio > 2.0:
            result.max_ltv = max(0, result.max_ltv - 0.05)
            result.warnings.append("ZHVI 200-300%: -5% leverage reduction")

    else:
        # RTL (Fix & Flip / Bridge / Bridge Plus)
        bclass = _eastview_rtl_classification(fico, exp)
        result.experience_tier = bclass
        result.fico_tier = f"FICO {fico}"

        # Determine product
        if is_bridge:
            grid = EV_RTL_BRIDGE
            product_name = "Bridge"
        elif deal.rehab_budget > 0:
            grid = EV_RTL_FF
            product_name = "Fix & Flip"
        else:
            grid = EV_RTL_BRIDGE
            product_name = "Bridge"

        purpose_data = grid.get(purpose, {})
        leverage = purpose_data.get(bclass)
        if leverage is None:
            result.eligible = False
            result.ineligible_reasons.append(f"Eastview RTL: Class {bclass} not eligible for {product_name} {purpose}")
            return result

        max_ltv, max_ltc, max_ltarv = leverage
        result.max_ltv = max_ltv or 0
        result.max_ltc = max_ltc or 0
        result.max_ltarv = max_ltarv or 0
        result.max_loan_amount = max_loan

        # Leverage reductions
        # Heavy rehab: budget > $50K AND > 100% of purchase price
        if deal.rehab_budget > 50000 and deal.purchase_price > 0:
            if deal.rehab_budget > deal.purchase_price:
                if bclass in ("A+", "A", "B"):
                    reduction = 0.05
                else:
                    reduction = 0.10
                result.max_ltv = max(0, result.max_ltv - reduction)
                if result.max_ltc:
                    result.max_ltc = max(0, result.max_ltc - reduction)
                if result.max_ltarv:
                    result.max_ltarv = max(0, result.max_ltarv - reduction)
                result.warnings.append(f"Heavy rehab: -{reduction:.0%} leverage reduction (Class {bclass})")

        # ZHVI multiplier reduction
        if deal.value_zhvi_ratio > 2.0:
            result.max_ltv = max(0, result.max_ltv - 0.05)
            result.warnings.append("ZHVI 200-300%: -5% leverage reduction")

        # Large loan reduction
        if deal.total_loan >= 2000000:
            result.max_ltv = max(0, result.max_ltv - 0.05)
            result.warnings.append("Loan $2M-$3M: -5% leverage reduction")

    # --- Check if deal fits within limits ---
    if result.max_ltv > 0 and deal.ltv > result.max_ltv:
        result.warnings.append(f"Deal LTV {deal.ltv:.1%} exceeds max {result.max_ltv:.1%}")
    if result.max_ltarv > 0 and deal.ltarv > result.max_ltarv:
        result.warnings.append(f"Deal LTARV {deal.ltarv:.1%} exceeds max {result.max_ltarv:.1%}")
    if result.max_ltc > 0 and deal.ltc > result.max_ltc:
        result.warnings.append(f"Deal LTC {deal.ltc:.1%} exceeds max {result.max_ltc:.1%}")

    # Rate — Eastview doesn't have a public rate sheet in guidelines
    result.estimated_rate = None
    result.rate_notes = "Use The Sizernator for full Eastview rate analysis"

    return result


# ---------------------------------------------------------------------------
# AI Recommendation Engine
# ---------------------------------------------------------------------------

def generate_recommendation(
    api_key: str,
    deal: DealParams,
    results: list[PartnerResult],
) -> str:
    """Use Claude to generate a recommendation comparing all partners."""
    client = anthropic.Anthropic(api_key=api_key)

    # Build deal summary
    deal_summary = f"""
Deal: {deal.deal_type} — {deal.transaction_type}
Property: {deal.address}, {deal.city}, {deal.state} {deal.zip_code}
Type: {deal.property_type} ({deal.num_units} units)
Purchase Price: ${deal.purchase_price:,.0f}
As-Is Value: ${deal.as_is_value:,.0f}
ARV: ${deal.arv:,.0f}
Rehab Budget: ${deal.rehab_budget:,.0f}
Requested Loan: ${deal.loan_amount:,.0f} + ${deal.rehab_holdback:,.0f} holdback + ${deal.interest_reserve:,.0f} reserves = ${deal.total_loan:,.0f}
LTV: {deal.ltv:.1%} | LTARV: {deal.ltarv:.1%} | LTC: {deal.ltc:.1%}
Borrower FICO: {deal.primary_fico} | Experience: {deal.completed_projects} projects
Liquidity: ${deal.verified_liquidity:,.0f}
ZHVI Ratio: {deal.value_zhvi_ratio:.1%}
"""

    # Build partner results summary
    partner_summaries = []
    for r in results:
        status = "ELIGIBLE" if r.eligible else "INELIGIBLE"
        summary = f"\n**{r.partner_name}** — {status}\n"
        if not r.eligible:
            summary += "Reasons: " + "; ".join(r.ineligible_reasons) + "\n"
        else:
            summary += f"Max LTV: {r.max_ltv:.1%} | Max LTARV: {r.max_ltarv:.1%} | Max LTC: {r.max_ltc:.1%}\n"
            if r.max_loan_amount > 0:
                summary += f"Max Loan: ${r.max_loan_amount:,.0f}\n"
            if r.estimated_rate is not None:
                summary += f"Estimated Rate: {r.estimated_rate:.3%}\n"
            if r.rate_notes:
                summary += f"Rate Notes: {r.rate_notes}\n"
            summary += f"Tier: {r.experience_tier} | FICO: {r.fico_tier}\n"
            if r.warnings:
                summary += "Warnings:\n" + "\n".join(f"  - {w}" for w in r.warnings) + "\n"
        partner_summaries.append(summary)

    prompt = f"""You are a senior loan analyst at A&S Capital, a private lending company.
Analyze this deal and the capital partner evaluations below. Provide a clear, concise recommendation.

{deal_summary}

Partner Evaluations:
{"".join(partner_summaries)}

Write a 3-5 sentence recommendation:
1. Which capital partner is the BEST fit for this deal and why
2. Key advantages of the recommended partner for this specific deal
3. Any concerns or conditions to be aware of
4. If multiple partners are eligible, briefly note the runner-up

Keep it professional, direct, and actionable. Do not use bullet points — write in paragraph form.
If no partners are eligible, explain why and suggest what would need to change.
"""

    response = client.messages.create(
        model="claude-4-sonnet-20250514",
        max_tokens=500,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_dealfit(
    api_key: str,
    file_bytes: bytes,
) -> tuple[DealParams, list[PartnerResult], str]:
    """
    Run the full Dealfit analysis.

    Returns: (deal_params, partner_results, ai_recommendation)
    """
    # 1. Read the sizer
    deal = read_sizer(file_bytes)

    # 2. Evaluate against each partner
    results = [
        evaluate_colchis(deal),
        evaluate_fidelis(deal),
        evaluate_eastview(deal),
    ]

    # 3. Generate AI recommendation
    recommendation = generate_recommendation(api_key, deal, results)

    return deal, results, recommendation
